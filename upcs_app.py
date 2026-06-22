import streamlit as st
import pandas as pd
import io
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pbi_downloader import run_download

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "template_upcs.xlsx")
SIZES = ["2T", "3T", "4", "4T", "5", "6", "6X", "7", "OS", "XS", "S", "M", "L", "XL", "2XL", "3XL"]

CURRENCY_MAP = {
    "USD": ("wholesale_usd", "msrp_usd"),
    "CAD": ("wholesale_cad", "msrp_cad"),
    "GBP": ("wholesale_gbp", "msrp_gbp"),
    "EUR": ("wholesale_eur", "msrp_eur"),
    "AED": ("wholesale_aed", "msrp_aed"),
    "MXN": ("wholesale_mxn", "msrp_mxn"),
    "BRL": ("wholesale_brl", "msrp_brl"),
    "CLP": ("wholesale_clp", "msrp_clp"),
    "AUD": ("wholesale_aud", "msrp_aud"),
    "NZD": ("wholesale_nzd", "msrp_nzd"),
    "RMB": ("wholesale_rmb", "msrp_rmb"),
    "ARS": ("wholesale_ars", "msrp_ars"),
    "ECU": ("wholesale_ecu", "msrp_ecu"),
    "BOB": ("wholesale_bob", "msrp_bob"),
    "PEN": ("wholesale_pen", "msrp_pen"),
}

BRAND_MAP = {
    "Pro Standard":    ["Pro Standard"],
    "Off White / L/AB": ["Off-White Division"],
}

st.set_page_config(page_title="UPCs Generator", page_icon="🏷️", layout="centered")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=DM+Sans:wght@300;400;500&display=swap');
html,body,[class*="css"]{font-family:'DM Sans',sans-serif;}
.main-title{font-family:'Bebas Neue',sans-serif;font-size:3rem;letter-spacing:4px;color:#e8c84a;}
.sub-title{font-family:'Bebas Neue',sans-serif;font-size:1.1rem;letter-spacing:3px;color:#888;}
.success-box{background:rgba(74,222,128,0.1);border:1px solid rgba(74,222,128,0.4);border-radius:6px;padding:12px 16px;font-size:0.9rem;color:#4ade80;}
.step-box{background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.1);border-radius:6px;padding:12px 16px;font-size:0.85rem;color:#ccc;margin-bottom:8px;}
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-title">UPCs GENERATOR</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">MAXIMA APPAREL</div>', unsafe_allow_html=True)
st.markdown("---")

st.markdown("#### Paso 1 — Obtén los datos")

tab_auto, tab_manual = st.tabs(["🔄 Automático (Base de datos)", "📁 Manual (subir archivo)"])

with tab_auto:
    st.markdown("""
    <div class="step-box">
    Conecta directamente a la base de datos de Maxima Apparel y descarga los UPCs automáticamente.
    </div>
    """, unsafe_allow_html=True)

    if "db_df" not in st.session_state:
        st.session_state.db_df = None
    if "db_log" not in st.session_state:
        st.session_state.db_log = []

    if st.button("🔌 Conectar y Descargar", use_container_width=True):
        st.session_state.db_df = None
        st.session_state.db_log = []
        st.session_state.pop("output", None)
        log_box = st.empty()

        def progress(msg):
            st.session_state.db_log.append(msg)
            log_box.markdown("\n\n".join(f"• {m}" for m in st.session_state.db_log))

        with st.spinner("Conectando..."):
            result = run_download(progress_fn=progress)
        st.session_state.db_df = result if result is not None and not isinstance(result, list) else None

    if st.session_state.db_log:
        with st.expander("Log de conexión", expanded=True):
            for msg in st.session_state.db_log:
                color = "#f87171" if "ERROR" in msg.upper() else "#4ade80" if "obtenidos" in msg.lower() else "#ccc"
                st.markdown(f'<div style="font-size:0.85rem;color:{color};">• {msg}</div>', unsafe_allow_html=True)

with tab_manual:
    st.markdown("""
    <div class="step-box">
    Exporta manualmente desde Power BI y sube el archivo aquí.
    </div>
    """, unsafe_allow_html=True)
    uploaded = st.file_uploader("Sube el archivo exportado de Power BI", type=["xlsx", "csv"], accept_multiple_files=True, key="file_upload")


def build_base_from_df(df, brands=None):
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]
    rename = {}
    for col in df.columns:
        if col in ("ivstyle", "style", "estilo"):
            rename[col] = "Style"
        elif col in ("size", "talla"):
            rename[col] = "Size"
        elif col == "upc":
            rename[col] = "UPC"
        elif col in ("description", "ivdesc", "descripcion", "descripción"):
            rename[col] = "DESCRIPTION"
        elif col in ("wholesale_usd", "wholesale", "mayoreo"):
            rename[col] = "WHOLESALE"
        elif col in ("msrp_usd", "msrp"):
            rename[col] = "MSRP"
        elif col in ("brand_name", "reporting_brand_name"):
            rename[col] = col
    df = df.rename(columns=rename)

    if brands:
        brand_map = {
            "Pro Standard": ["Pro Standard"],
            "Off White / L/AB": ["Off-White Division"],
        }
        allowed = []
        for b in brands:
            allowed.extend(brand_map.get(b, [b]))
        if "reporting_brand_name" in df.columns:
            df = df[df["reporting_brand_name"].astype(str).str.strip().isin(allowed)]

    needed = [c for c in ["Style", "Size", "UPC", "DESCRIPTION", "WHOLESALE", "MSRP"] if c in df.columns]
    df = df[needed].copy()
    df["UPC"] = df["UPC"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    df = df[df["UPC"].str.len() > 3]
    df = df[df["UPC"].str.lower() != "nan"]
    if "Size" in df.columns:
        df = df[df["Size"].astype(str).str.strip().isin(SIZES)]
    df = df[df["Style"].notna() & (df["Style"].astype(str).str.strip() != "")]
    df = df.drop_duplicates(subset=["Style", "Size"], keep="first")
    return df.reset_index(drop=True)


def build_base_from_files(files):
    all_rows = []
    for f in files:
        for header_row in [2, 0, 1, 3]:
            try:
                df = pd.read_excel(f, header=header_row, dtype=str)
                df.columns = [str(c).strip() for c in df.columns]
                rename = {}
                for col in df.columns:
                    cl = col.lower().strip()
                    if cl in ["ivnum", "estilo", "style"]:
                        rename[col] = "Style"
                    elif cl in ["size", "talla"]:
                        rename[col] = "Size"
                    elif cl == "upc":
                        rename[col] = "UPC"
                    elif cl in ["description", "descripcion", "descripción"]:
                        rename[col] = "DESCRIPTION"
                    elif cl in ["wholesale", "mayoreo"]:
                        rename[col] = "WHOLESALE"
                    elif cl == "msrp":
                        rename[col] = "MSRP"
                df = df.rename(columns=rename)
                if "Style" not in df.columns or "UPC" not in df.columns:
                    continue
                needed = [c for c in ["Style", "Size", "UPC", "DESCRIPTION", "WHOLESALE", "MSRP"] if c in df.columns]
                df = df[needed].copy()
                df["UPC"] = df["UPC"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
                df = df[df["UPC"].str.strip().str.len() > 3]
                df = df[df["UPC"].str.strip() != "nan"]
                if "Size" in df.columns:
                    df = df[df["Size"].isin(SIZES)]
                df = df[df["Style"].notna() & (df["Style"].str.strip() != "")]
                all_rows.append(df)
                break
            except Exception:
                continue

    if not all_rows:
        return pd.DataFrame(columns=["Style", "Size", "UPC", "DESCRIPTION", "WHOLESALE", "MSRP"])
    base = pd.concat(all_rows, ignore_index=True)
    if "Style" in base.columns and "Size" in base.columns:
        base = base.drop_duplicates(subset=["Style", "Size"], keep="first")
    return base.reset_index(drop=True)


def _sheet_xml_path(zf, sheet_name):
    """Devuelve el path interno (xl/worksheets/sheetN.xml) para un nombre de hoja."""
    NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    R  = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    wb_root = ET.fromstring(zf.read('xl/workbook.xml'))
    rid = next(
        (s.get(f'{{{R}}}id') for s in wb_root.iter(f'{{{NS}}}sheet')
         if s.get('name') == sheet_name), None)
    if not rid:
        raise ValueError(f"Hoja '{sheet_name}' no encontrada en el workbook.")
    rels = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
    target = next((r.get('Target') for r in rels if r.get('Id') == rid), None)
    if not target:
        raise ValueError(f"No se encontró el archivo para la hoja '{sheet_name}'.")
    return 'xl/' + target


def fill_template_base(base_df):
    """
    Llena la hoja BASE del template con los datos del DataFrame.

    Usa openpyxl para escribir los datos y zipfile para restaurar
    los archivos internos (queryTables, rels) que openpyxl elimina
    al guardar, evitando el error 'Removed Records: Formula from sheet3.xml'.
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(
            f"No se encontró template_upcs.xlsx en:\n{TEMPLATE_PATH}\n"
            "Colócalo en la misma carpeta que upcs_app.py."
        )

    MONEY_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

    # ── 1. Llenar datos con openpyxl ─────────────────────────────────────
    wb = load_workbook(TEMPLATE_PATH)
    if "BASE" not in wb.sheetnames:
        raise ValueError("El template no tiene hoja 'BASE'.")
    ws = wb["BASE"]

    # Leer encabezados
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip().upper()] = cell.column

    # Borrar filas de datos (conservar encabezado)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # Mapeo de columnas
    col_aliases = {
        "Style": ["STYLE"], "Size": ["SIZE"], "UPC": ["UPC"],
        "DESCRIPTION": ["DESCRIPTION"],
        "WHOLESALE": ["WHOLESALE", "WHOLESAL"], "MSRP": ["MSRP"],
    }
    fill_map = {}
    for df_col, names in col_aliases.items():
        for n in names:
            if n in headers:
                fill_map[df_col] = headers[n]
                break

    # Escribir datos
    for r_idx, (_, row) in enumerate(base_df.iterrows(), start=2):
        for df_col, col_idx in fill_map.items():
            if df_col not in base_df.columns:
                continue
            value = row[df_col]
            try:
                if pd.isna(value):
                    value = None
            except (TypeError, ValueError):
                pass
            cell = ws.cell(row=r_idx, column=col_idx, value=value)
            if df_col == "UPC":
                cell.number_format = "@"
            elif df_col in ("WHOLESALE", "MSRP"):
                cell.number_format = MONEY_FMT

    # Actualizar ref de tablas Excel si existen (evita rango obsoleto)
    for tbl in ws.tables.values():
        tbl.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    pl_buf = io.BytesIO()
    wb.save(pl_buf)
    pl_buf.seek(0)

    # ── 2. Recolectar lo que openpyxl pudo haber eliminado del template ──
    with zipfile.ZipFile(TEMPLATE_PATH, 'r') as tmpl:
        tmpl_names = set(tmpl.namelist())
        base_path_tmpl = _sheet_xml_path(tmpl, 'BASE')
        base_file_tmpl = base_path_tmpl.split('/')[-1]
        rels_path_tmpl = f'xl/worksheets/_rels/{base_file_tmpl}.rels'

        to_restore = {}
        if rels_path_tmpl in tmpl_names:
            to_restore[rels_path_tmpl] = tmpl.read(rels_path_tmpl)
        for f in tmpl_names:
            if re.search(r'queryTable', f, re.IGNORECASE):
                to_restore[f] = tmpl.read(f)

        # Bloque <queryTableParts> de la hoja BASE original
        orig_xml = tmpl.read(base_path_tmpl).decode('utf-8')
        qt_match = re.search(
            r'<queryTableParts\b(?:[^>]*/?>|.*?</queryTableParts>)',
            orig_xml, re.DOTALL)
        qt_block = qt_match.group(0) if qt_match else None

        # Entradas de Content_Types para queryTable
        ct_orig = tmpl.read('[Content_Types].xml').decode('utf-8')
        qt_ct = re.findall(r'<(?:Override|Default)[^>]+[Qq]uery[^>]+/>', ct_orig)

    # ── 3. Reconstruir zip con archivos restaurados ───────────────────────
    output_buf = io.BytesIO()
    with zipfile.ZipFile(pl_buf, 'r') as pl:
        pl_names = set(pl.namelist())
        new_base_path = _sheet_xml_path(pl, 'BASE')
        new_base_file = new_base_path.split('/')[-1]
        new_rels_path = f'xl/worksheets/_rels/{new_base_file}.rels'

        with zipfile.ZipFile(output_buf, 'w', zipfile.ZIP_DEFLATED) as out:
            for item in pl.namelist():
                data = pl.read(item)

                # Re-inyectar <queryTableParts> en la hoja BASE
                if item == new_base_path and qt_block:
                    s = data.decode('utf-8')
                    if '<queryTableParts' not in s:
                        s = s.replace('</worksheet>', f'{qt_block}</worksheet>')
                    data = s.encode('utf-8')

                # Añadir entradas de Content_Types para queryTable
                if item == '[Content_Types].xml' and qt_ct:
                    s = data.decode('utf-8')
                    for entry in qt_ct:
                        if entry not in s:
                            s = s.replace('</Types>', f'{entry}</Types>')
                    data = s.encode('utf-8')

                out.writestr(item, data)

            # Restaurar archivos que openpyxl eliminó
            for path, content in to_restore.items():
                mapped = path.replace(
                    f'_rels/{base_file_tmpl}.rels',
                    f'_rels/{new_base_file}.rels'
                )
                if mapped not in pl_names:
                    out.writestr(mapped, content)

    output_buf.seek(0)
    return output_buf.getvalue()


db_df = st.session_state.get("db_df", None)
has_data = db_df is not None or (uploaded and len(uploaded) > 0)

st.markdown("#### Paso 2 — Genera el archivo")
tab_base, tab_search = st.tabs(["📊 BASE Completo", "🔍 Búsqueda por Estilo / Moneda"])

# ── TAB 1: BASE Completo ─────────────────────────────────────────────────
with tab_base:
    file_date = st.date_input("Fecha", value=date.today(), key="date_base")

    BRANDS = ["Pro Standard", "Off White / L/AB"]
    selected_brands = st.multiselect(
        "Filtrar por marca (vacío = todas)",
        options=BRANDS, default=[],
        placeholder="Selecciona una o más marcas...",
        key="brands_base",
    )

    if has_data:
        if st.button("GENERAR BASE", type="primary", use_container_width=True):
            st.session_state.pop("output", None)
            with st.spinner("Procesando..."):
                if db_df is not None:
                    base_df = build_base_from_df(db_df, brands=selected_brands)
                else:
                    base_df = build_base_from_files(list(uploaded) if uploaded else [])

            if base_df.empty:
                st.error("No se encontraron datos.")
            else:
                try:
                    xlsx_bytes = fill_template_base(base_df)
                    fname = f"UPCS_{file_date.strftime('%m.%d.%Y')}.xlsx"
                    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "UPCs_generados")
                    os.makedirs(out_dir, exist_ok=True)
                    out_path = os.path.join(out_dir, fname)
                    try:
                        with open(out_path, "wb") as fh:
                            fh.write(xlsx_bytes)
                        saved_path = out_path
                    except Exception:
                        saved_path = None
                    st.session_state["output"] = {
                        "fname": fname, "bytes": xlsx_bytes,
                        "count": len(base_df), "path": saved_path,
                    }
                except FileNotFoundError as e:
                    st.error(str(e))
                except Exception as e:
                    st.error(f"Error al generar: {e}")
    else:
        st.info("Primero conecta a la base de datos (Paso 1).")

    out = st.session_state.get("output")
    if out:
        st.markdown(f'<div class="success-box">✅ Listo — {out["count"]:,} registros en BASE</div>', unsafe_allow_html=True)
        if out.get("path"):
            st.success(f"Guardado en:\n\n`{out['path']}`\n\nAbre el archivo → escribe estilos en **Styles** → **Datos → Actualizar todo**")
        st.download_button(
            label=f"⬇ Descargar {out['fname']}",
            data=out["bytes"], file_name=out["fname"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

# ── TAB 2: Búsqueda por Estilo / Moneda ─────────────────────────────────
with tab_search:
    st.markdown("Busca estilos específicos y selecciona la moneda para los precios.")

    col1, col2 = st.columns(2)

    with col1:
        styles_input = st.text_area(
            "Estilos (uno por línea)",
            placeholder="BCB1517653-WHT\nFPE6411013-KGR\n...",
            height=180,
            key="styles_input",
        )

    with col2:
        # Marcas disponibles (dinámico si hay datos cargados)
        if db_df is not None and "reporting_brand_name" in db_df.columns:
            available_brands = sorted(db_df["reporting_brand_name"].dropna().unique().tolist())
        else:
            available_brands = list(BRAND_MAP.keys())

        selected_brands_search = st.multiselect(
            "Marca (vacío = todas)",
            options=available_brands, default=[],
            placeholder="Selecciona marca(s)...",
            key="brands_search",
        )

        selected_currencies = st.multiselect(
            "Moneda(s)",
            options=list(CURRENCY_MAP.keys()),
            default=["USD"],
            placeholder="Selecciona moneda(s)...",
            key="currencies_search",
        )

    file_date_search = st.date_input("Fecha", value=date.today(), key="date_search")

    if st.button("🔍 BUSCAR Y GENERAR", type="primary", use_container_width=True):
        st.session_state.pop("output_search", None)

        if db_df is None:
            st.error("Primero conecta a la base de datos (Paso 1).")
        elif not styles_input.strip():
            st.error("Escribe al menos un estilo.")
        elif not selected_currencies:
            st.error("Selecciona al menos una moneda.")
        else:
            styles_list = [s.strip() for s in styles_input.strip().splitlines() if s.strip()]

            with st.spinner("Buscando..."):
                df = db_df.copy()
                df.columns = [str(c).strip().lower() for c in df.columns]

                # Filtrar por estilos
                style_col = next((c for c in df.columns if c in ("style", "ivnum")), None)
                if style_col:
                    df = df[df[style_col].astype(str).str.strip().isin(styles_list)]

                # Filtrar por marca
                if selected_brands_search and "reporting_brand_name" in df.columns:
                    # Expandir alias si son marcas del brand_map
                    allowed = []
                    for b in selected_brands_search:
                        allowed.extend(BRAND_MAP.get(b, [b]))
                    df = df[df["reporting_brand_name"].astype(str).str.strip().isin(allowed)]

                # Filtrar por tallas
                if "size" in df.columns:
                    df = df[df["size"].astype(str).str.strip().isin(SIZES)]

                # Limpiar UPC
                if "upc" in df.columns:
                    df["upc"] = df["upc"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
                    df = df[df["upc"].str.len() > 3]
                    df = df[df["upc"].str.lower() != "nan"]

                # Quitar duplicados Style+Size
                if style_col and "size" in df.columns:
                    df = df.drop_duplicates(subset=[style_col, "size"], keep="first")

            if df.empty:
                st.warning("No se encontraron estilos. Verifica que estén en la base de datos y que los datos estén descargados.")
            else:
                # Construir columnas del Excel
                HEADER_FILL = PatternFill("solid", fgColor="006699")
                HEADER_FONT = Font(bold=True, color="FFFFFF")
                MONEY_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                wb = Workbook()
                ws = wb.active
                ws.title = "UPCs"

                # Headers
                headers = ["Styles", "Size", "UPC", "DESCRIPTION"]
                for curr in selected_currencies:
                    headers.append(f"WholeSale_{curr}")
                    headers.append(f"MSRP_{curr}")
                ws.append(headers)

                # Data
                desc_col = next((c for c in df.columns if c in ("description", "ivdesc")), None)
                for _, row in df.iterrows():
                    style_val = str(row.get(style_col, "") or "")
                    size_val  = str(row.get("size", "") or "")
                    upc_val   = str(row.get("upc", "") or "")
                    desc_val  = str(row.get(desc_col, "") or "") if desc_col else ""

                    data_row = [style_val, size_val, upc_val, desc_val]
                    for curr in selected_currencies:
                        ws_col, ms_col = CURRENCY_MAP[curr]
                        ws_val = row.get(ws_col)
                        ms_val = row.get(ms_col)
                        try:
                            ws_val = float(ws_val) if ws_val is not None and str(ws_val) not in ("", "nan", "None") else None
                        except (ValueError, TypeError):
                            ws_val = None
                        try:
                            ms_val = float(ms_val) if ms_val is not None and str(ms_val) not in ("", "nan", "None") else None
                        except (ValueError, TypeError):
                            ms_val = None
                        data_row.extend([ws_val, ms_val])
                    ws.append(data_row)

                # Format UPC column as text
                upc_col_idx = 3
                for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=upc_col_idx, max_col=upc_col_idx):
                    cell[0].number_format = "@"

                # Format price columns
                price_start = 5
                for c_idx in range(price_start, price_start + len(selected_currencies) * 2):
                    for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=c_idx, max_col=c_idx):
                        cell[0].number_format = MONEY_FMT

                # Style header
                for cell in ws[1]:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.auto_filter.ref = ws.dimensions
                ws.freeze_panes = "A2"

                # Autofit columns
                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=0)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 50)

                buf = io.BytesIO()
                wb.save(buf)
                xlsx_bytes = buf.getvalue()

                curr_label = "_".join(selected_currencies)
                fname_s = f"UPCS_Estilos_{curr_label}_{file_date_search.strftime('%m.%d.%Y')}.xlsx"

                out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "UPCs_generados")
                os.makedirs(out_dir, exist_ok=True)
                out_path = os.path.join(out_dir, fname_s)
                try:
                    with open(out_path, "wb") as fh:
                        fh.write(xlsx_bytes)
                    saved_path_s = out_path
                except Exception:
                    saved_path_s = None

                st.session_state["output_search"] = {
                    "fname": fname_s, "bytes": xlsx_bytes,
                    "count": len(df), "path": saved_path_s,
                }

    out_s = st.session_state.get("output_search")
    if out_s:
        st.markdown(f'<div class="success-box">✅ {out_s["count"]:,} registros encontrados</div>', unsafe_allow_html=True)
        if out_s.get("path"):
            st.success(f"Guardado en:\n\n`{out_s['path']}`")
        st.download_button(
            label=f"⬇ Descargar {out_s['fname']}",
            data=out_s["bytes"], file_name=out_s["fname"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="dl_search",
        )
