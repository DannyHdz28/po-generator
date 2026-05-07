import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import re
import io
from datetime import datetime, date

# ─── PAGE CONFIG ────────────────────────────────────────────
st.set_page_config(page_title="Pro Standard — PO Generator", page_icon="🏆", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=DM+Sans:wght@300;400;500&display=swap');
html,body,[class*="css"]{font-family:'DM Sans',sans-serif;}
.main-title{font-family:'Bebas Neue',sans-serif;font-size:3rem;letter-spacing:4px;color:#e8c84a;}
.sub-title{font-family:'Bebas Neue',sans-serif;font-size:1.2rem;letter-spacing:3px;color:#888;}
.section-title{font-family:'Bebas Neue',sans-serif;font-size:1.1rem;letter-spacing:2px;color:#aaa;}
.field-box{background:#1a1a1a;border-radius:6px;padding:10px 14px;margin:4px 0;}
.field-label{font-size:0.7rem;letter-spacing:1.5px;text-transform:uppercase;color:#666;}
.field-value{font-size:0.95rem;color:#f5f5f0;}
.warn-box{background:rgba(255,165,0,0.1);border:1px solid rgba(255,165,0,0.4);border-radius:6px;padding:10px 14px;font-size:0.85rem;color:#ffa500;}
.success-box{background:rgba(74,222,128,0.1);border:1px solid rgba(74,222,128,0.4);border-radius:6px;padding:10px 14px;font-size:0.85rem;color:#4ade80;}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# GLOSARIO COMPLETO
# Basado en Glosario_PO.docx + POs reales: Kings, Fitters, Fanatics, Dallas Cowboys, Pro Standard HO
# ══════════════════════════════════════════════════════════════
GLOSSARY = {
    # ── Encabezado ────────────────────────────────────────────
    "po_number": [
        "PO NUMBER","PO NO","PO NO.","PO NAME",
        "PURCHASE ORDER NUMBER","ORDER NUMBER",
        "P.O. NUMBER","P.O.#",
        "CUSTOMER PO",  # Boom format
        # Short/ambiguous removed: "PO#","PO:","ORDER #","ORDER NO" -- cause false positives
    ],
    "po_date": [
        "PO DATE","ORDER DATE","PO CREATION DATE","CREATION DATE",
        "ISSUE DATE","CREATED ON","DATE ISSUED","DATE",
        "FECHA DE EMISION","FECHA DE LA ORDEN","FECHA DE GENERACION",
    ],
    "ship_date": [
        # Glosario: Start Ship, Ship Window Open, In-DC Date, Delivery Date, Dispatch Date
        # Fitters usa "Start Date", Kings usa "Ship Date", Dallas Cowboys usa "Delivery Date"
        "SHIP DATE","START DATE","START SHIP","START SHIP DATE","SHIP WINDOW OPEN",
        "IN-DC DATE","IN DC DATE","DELIVERY DATE","DISPATCH DATE",
        "SHIP BY","DELIVER BY","IN HANDS DATE","IN-HANDS DATE",
        "IN HANDS","NEED BY DATE","ON FLOOR",
        "FECHA DE EMBARQUE","FECHA DE ENVIO","FECHA PROGRAMADA",
    ],
    "cancel_date": [
        # Glosario: Cancel-by Date, DNSA, Expiration Date, Latest Ship Date
        "CANCEL DATE","CANCEL BY","CANCEL-BY DATE","CANCELLATION DATE",
        "TERMINATION DATE","DO NOT SHIP AFTER","DNSA",
        "EXPIRATION DATE","LATEST SHIP DATE",
        "FECHA LIMITE","FECHA DE CANCELACION","FECHA TOPE",
    ],
    "customer_name": [
        # Glosario: Buyer, Bill-to Name, Account Name, Cliente
        # NOTE: "BILL TO" handled by special block below — NOT here
        # Boom usa "Bill To Name" as a label:value pair
        "CUSTOMER NAME","SOLD TO","ACCOUNT NAME",
        "BILLED TO","CLIENT","BILL TO NAME",
        "RAZON SOCIAL","COMPRADOR","CONSIGNATARIO",
    ],
    "customer_code": [
        # Glosario: Account Number, Account #, Customer ID, Buyer Code
        "CUSTOMER CODE","ACCOUNT NUMBER","ACCOUNT #","ACCOUNT NO",
        "CUSTOMER ID","BUYER CODE","CLIENT CODE",
        "CUENTA DEL CLIENTE","ID DE CLIENTE","NO. DE CUENTA",
    ],
    "ship_to": [
        # Glosario: Delivery Address, Consignee, DC Address, Store Address
        "SHIP TO","SHIP-TO","SHIPPING ADDRESS","DELIVER TO",
        "DELIVERY ADDRESS","CONSIGNEE","DC ADDRESS","STORE ADDRESS",
        "DROP ADDRESS","DIRECCION DE ENVIO","DESTINATARIO",
    ],
    "bill_to": [
        # Glosario: Invoice Address, Sold to, Accounts Payable Address
        "BILL TO","BILL-TO","INVOICE ADDRESS","ACCOUNTS PAYABLE",
        "AP ADDRESS","BILLING ADDRESS","INVOICE TO","FACTURAR A",
    ],
    "terms": ["TERMS","PAYMENT TERMS","NET TERMS","NET"],

    # ── Columnas de tabla ─────────────────────────────────────
    "col_style": [
        # Glosario: Style Number, Style Code, Vendor Item Number, Model Number
        # Visto: Vendor Style No. (Kings), Style Number (Fitters), Vendor Item Number (Caesars)
        "VENDOR STYLE NO.","VENDOR STYLE NO","VENDOR STYLE NUMBER",
        "VENDOR STYLE#","VENDOR STYLE #","VENDOR STYLE","STYLE ID",
        "STYLE #","STYLE NO","STYLE NO.","STYLE NUMBER","STYLE CODE",
        "STOCK","ITEM #","ITEM NUMBER","ITEM NO","SKU","SKU#",
        "PRODUCT CODE","STYLE","VENDOR ITEM NUMBER","VENDOR ITEM",
        "ARTICLE","MODEL NUMBER","MANUFACTURER CODE",
        "ESTILO","REFERENCIA DEL PROVEEDOR",
    ],
    "col_desc": [
        # Glosario: Item Description, Product Name, Article Description
        "ITEM DESCRIPTION","ITEM DESC","DESCRIPTION","DESC",
        "PRODUCT NAME","PRODUCT TITLE","PRODUCT DESCRIPTION","ARTICLE DESCRIPTION",
        "NAME",  # Boom format uses "name" column for product description
        "DETALLE DEL PRODUCTO","DESCRIPCION DEL ARTICULO",
    ],
    "col_color": [
        # Glosario: Color, Colorway, Hue, Color ID, Color Name
        "COLOR NAME","COLOR STORY","COLOUR NAME","COLOR",
        "COLOUR","COLORWAY","HUE","COLOR ID","CODIGO DE COLOR",
    ],
    "col_size_break": [
        # Glosario: Size scale, Size run, Size distribution, Pack ratio
        # Visto en Pro Standard HO: "SIZE SCALE"
        "SIZE SCALE","SIZE BREAK","SIZE RUN","SIZE DISTRIBUTION",
        "PACK RATIO","SIZE CURVE","SIZES",
        "QUEBRA DE TALLAS","CURVA DE TALLAS","DISTRIBUCION POR TALLA",
    ],
    "col_qty": [
        # Glosario: Total Qty, Total Pieces, Total Units, Sum of Qty
        "QTY","QUANTITY","PURCH QTY","ORDER QTY",
        "UNITS","# UNITS","PIECES","PCS",
        "TTL UNITS","TTL QTY","TOTAL PIECES","SUM OF QTY",
    ],
    "col_cost": [
        # Glosario: Net cost, Unit cost, Price Per Unit, Wholesale net
        # Priority: net/discounted cost first, then gross cost
        # Visto: "Cost W/ Disc" (Kings), "Cost W/DISC." (Black Knights),
        #        "Price Per Unit" (Fitters), "Unit Cost" (Caesars), "Cost" (Dallas Cowboys)
        "COST W/ DISC","COST W/DISC","COST W/DISC.","COST WITH DISCOUNT",
        "NET COST","NET UNIT COST","UNIT COST","UNIT PRICE",
        "PRICE PER UNIT","WHOLESALE NET","DISCOUNTED PRICE",
        "SALE PRICE",  # Boom format
        "LINE COST",  # Maxima internal template (Borregos, Guerreros, Dicass)
        "WHOLESALE","COST","WS","PRICE",
        "COSTO NETO","PRECIO NETO","COSTO UNITARIO","MAYOREO NETO",
    ],
    "col_msrp": [
        # Glosario: Suggested Retail Price, Retail Price, List Price, Tag Price
        # Visto: "MSRP / Retail" (Fitters), "Retail Price" (Caesars), "Retail" (Kings)
        "MSRP","MSRP / RETAIL","RETAIL","RETAIL PRICE",
        "SUGGESTED RETAIL PRICE","SRP","LIST PRICE",
        "TAG PRICE","TICKET PRICE","PRECIO SUGERIDO","PVP",
    ],
    "col_total_cost": [
        # Glosario: Extended Cost, Ext Cost, Line Total, Total Dollars
        "TOTAL DOLLARS","EXT COST","EXTENDED COST","LINE TOTAL",
        "TOTAL COST","SUBTOTAL","LINE TOTAL","TOT. BUY","TOTAL NET COST",
        "COSTO EXTENDIDO","COSTO POR RENGLON",
    ],
    "col_total_retail": [
        # Glosario: Ext Retail, Retail Total, Sales Value
        "EXT RETAIL","EXTENDED RETAIL","RETAIL VALUE",
        "TOTAL RETAIL","RETAIL TOTAL","SALES VALUE","VALOR DE VENTA",
    ],
    "col_discount": [
        # Only pure percentage/rate columns — NOT cost columns
        "DISC %","DISCOUNT %","DISC","DISCOUNT","OFF %","DESCUENTO",
    ],
    "col_upc": [
        # Glosario: Barcode, GTIN, EAN, Item barcode, Scan code
        "UPC","UPC CODE","BARCODE","GTIN","EAN",
        "EAN-13","SCAN CODE","ITEM BARCODE","CODIGO DE BARRAS",
    ],

    # ── Tallas individuales (formato matriz / columnas separadas) ──
    # Glosario: "Puede aparecer como matriz horizontal (XS,S,M,L,XL,2X,3X)"
    # Visto en Fitters y Saleh Sportswear
    "size_os":  ["OS","OSFA","OSFM","ONE SIZE","O/S","ONE-SIZE","ONE SIZE FITS ALL"],
    "size_xxs": ["XXS","2XS"],
    "size_xs":  ["XS"],
    "size_s":   ["S","SM","SML","S/P","SMALL"],
    "size_m":   ["M","MD","MED","M/M","MEDIUM"],
    "size_l":   ["L","LG","LRG","L/G","LARGE"],
    "size_xl":  ["XL","X-LARGE","X/L","XLARGE"],
    "size_2xl": ["2XL","XXL","2X","XX","2X-LARGE","2X LARGE","DOUBLE XL"],
    "size_3xl": ["3XL","XXXL","3X","3X-LARGE","3X LARGE","TRIPLE XL"],
    "size_4xl": ["4XL","XXXXL","4X","4X-LARGE","4X LARGE"],
    "size_5xl": ["5XL","XXXXXL","5X","5X-LARGE","5X LARGE"],
    "size_4xl": ["4XL","XXXXL","4X","4X-LARGE","4X LARGE"],
    "size_5xl": ["5XL","XXXXXL","5X","5X-LARGE","5X LARGE"],
}

SIZE_ORDER   = ["OS","XXS","XS","S","M","L","XL","2XL","3XL","4XL","5XL"]
SIZE_KEY_MAP = {
    "size_os":"OS","size_xxs":"XXS","size_xs":"XS",
    "size_s":"S","size_m":"M","size_l":"L",
    "size_xl":"XL","size_2xl":"2XL","size_3xl":"3XL",
    "size_4xl":"4XL","size_5xl":"5XL",
}
SIZE_NORMALIZE = {
    "OSFA":"OS","OSFM":"OS","ONE SIZE":"OS","O/S":"OS","ONE-SIZE":"OS",
    "XXS":"XXS","2XS":"XXS",
    "XS":"XS",
    "S":"S","SM":"S","SML":"S","S/P":"S","SMALL":"S",
    "M":"M","MD":"M","MED":"M","M/M":"M","MEDIUM":"M",
    "L":"L","LG":"L","LRG":"L","L/G":"L","LARGE":"L",
    "XL":"XL","X-LARGE":"XL","X/L":"XL","XLARGE":"XL","EXTRA LARGE":"XL",
    "2XL":"2XL","XXL":"2XL","2X":"2XL","XX":"2XL","2X-LARGE":"2XL","2X LARGE":"2XL",
    "3XL":"3XL","XXXL":"3XL","3X":"3XL","3X-LARGE":"3XL","3X LARGE":"3XL",
    "4XL":"4XL","XXXXL":"4XL","4X":"4XL","4X-LARGE":"4XL","4X LARGE":"4XL",
    "5XL":"5XL","XXXXXL":"5XL","5X":"5XL","5X-LARGE":"5XL","5X LARGE":"5XL",
    "MD":"M","LG":"L","SM":"S",  # Caesars format
}
COLOR_CODES = {
    "BLK":"BLACK","WHT":"WHITE","NVY":"NAVY","RED":"RED","BLU":"BLUE",
    "GRY":"GREY","WBK":"WASHED BLACK","EGG":"EGGSHELL","CRM":"CREAM",
    "DBL":"DODGER BLUE","HGR":"HEATHER GREY","DHG":"DARK HEATHER GREY",
    "LGY":"LIGHT GREY","MDN":"MIDNIGHT NAVY","RYB":"ROYAL BLUE",
    "NSF":"NIGHT SKY FADE","GLD":"GOLD","PUR":"PURPLE","FOR":"FOREST",
    "TEL":"TEAL","KGN":"KELLY GREEN","CRD":"CRIMSON RED","CNC":"CHARCOAL",
}

# ══════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════

def cell_str(v):
    if v is None: return ""
    s = str(v).strip().lstrip("'")  # strip Excel text prefix apostrophe
    return s.upper().strip()

def matches_any(text, keywords):
    t = text.upper().strip().rstrip(":")  # strip trailing colon for label matching
    for kw in keywords:
        k = kw.upper().strip().rstrip(":")
        if not k: continue
        if t == k or t.startswith(k) or k in t:
            return True
    return False

def detect_col(header_row, keywords, exact_only=False):
    """Find column index. Prefers exact match over contains match.
    Among exact matches, prefers the one matching the LONGEST keyword (more specific).
    If exact_only=True, only returns exact matches (used for size columns)."""
    best_exact = None       # (col_index, keyword_length)
    best_contains = None    # (col_index, keyword_length)
    for i, cell in enumerate(header_row):
        c = cell_str(cell)
        if not c:
            continue
        for kw in keywords:
            k = kw.upper().strip()
            if c == k:
                if best_exact is None or len(k) > best_exact[1]:
                    best_exact = (i, len(k))
            elif not exact_only and k in c:
                if best_contains is None or len(k) > best_contains[1]:
                    best_contains = (i, len(k))
    if best_exact is not None:
        return best_exact[0]
    if not exact_only and best_contains is not None:
        return best_contains[0]
    return None

def normalize_size(s):
    u = str(s).upper().strip()
    return SIZE_NORMALIZE.get(u, u)

def fmtDate(v):
    if not v:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime('%m/%d/%Y')
    s = str(v).strip().lstrip("'")
    # Single char or very short — not a date
    if len(s) <= 2:
        return ""
    # Remove time portion
    s = re.sub(r'\s+\d{2}:\d{2}:\d{2}.*', '', s)
    # Convert ISO format 2026-10-01 -> 10/01/2026
    iso = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
    if iso:
        return f"{iso.group(2)}/{iso.group(3)}/{iso.group(1)}"
    # Clean double slashes
    s = re.sub(r'/+', '/', s)
    return s

def parse_style_color(style_code):
    s = str(style_code).strip()
    if "-" in s:
        parts = s.rsplit("-", 1)
        stock = parts[0].strip()
        cc    = parts[1].strip().upper()
        cn    = COLOR_CODES.get(cc, cc)
        return stock, cc, cn
    return s, "", ""

def parse_size_break_compressed(cell_value):
    sizes = {s: 0 for s in SIZE_ORDER}
    if not cell_value:
        return sizes
    val = str(cell_value).strip()
    parts = re.split(r'\s{2,}|\t', val, maxsplit=1)
    range_part   = parts[0].strip()
    numbers_part = parts[1].strip() if len(parts) > 1 else val
    range_match  = re.match(r'([A-Z0-9/\-]+?)\s*[-]+\s*([A-Z0-9]+)', range_part.upper())
    if range_match:
        start = normalize_size(range_match.group(1).split('/')[0].strip())
        end   = normalize_size(range_match.group(2).strip())
        try:
            si = SIZE_ORDER.index(start)
            ei = SIZE_ORDER.index(end)
            size_slice = SIZE_ORDER[si:ei+1]
        except ValueError:
            size_slice = ["S","M","L","XL","2XL","3XL"]
    else:
        size_slice = ["S","M","L","XL","2XL","3XL"]
    numbers = [int(n) for n in re.findall(r'\d+', numbers_part)]
    for i, sz in enumerate(size_slice):
        if i < len(numbers):
            sizes[sz] = numbers[i]
    return sizes

# Labels that should never be returned as values
_LABEL_WORDS = {
    "PO NUMBER","PO NAME","PO DATE","PO#","SHIP DATE","CANCEL DATE",
    "BILL TO","SHIP TO","VENDOR","BRAND","BUYER","TERMS","CURRENCY",
    "DESCRIPTION","STYLE","QTY","COST","RETAIL","MSRP","UPC",
    "UNITS","AMOUNT","START DATE","PO NAME","BORREGOS","CIMACO",
    "VENDOR:","TERMS:","CREDIT CARD","NET 30","NET 45","NET 60",
}

def find_header_raw(rows, keywords, max_rows=25):
    for row in rows[:max_rows]:
        for j, cell in enumerate(row):
            c = cell_str(cell)
            if c and matches_any(c, keywords):
                raw = str(cell).strip()
                if ":" in raw:
                    after = raw.split(":", 1)[1].strip()
                    if after and after not in ("", "0", "None"):
                        return after
                for k in range(j+1, min(j+5, len(row))):
                    v = row[k]
                    if v is None:
                        continue
                    vs = str(v).strip().lstrip("'")
                    if vs in ("", "0", "None"):
                        continue
                    # Skip values that look like column labels
                    vs_upper = vs.upper()
                    is_label = (vs_upper in _LABEL_WORDS or 
                                any(vs_upper.startswith(lw) for lw in ["TERMS:", "VENDOR:", "SHIP TO:", "BILL TO:", "CANCEL", "NET "]))
                    if is_label:
                        continue
                    # Return original value for dates (preserve datetime objects)
                    if isinstance(v, (datetime, date)):
                        return v
                    return vs
    return None

def find_header_value(rows, keywords, max_rows=25):
    v = find_header_raw(rows, keywords, max_rows)
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v
    return str(v).strip()


# ══════════════════════════════════════════════════════════════
# PDF PARSER — uses pdfplumber, no API key needed
# Works for PDFs with real text (not scanned images)
# ══════════════════════════════════════════════════════════════

def parse_po_pdf(file_bytes):
    """Parse a PDF purchase order using pdfplumber."""
    try:
        import pdfplumber
    except ImportError:
        return {"po_number":"","po_date":"","ship_date":"","cancel_date":"",
                "customer_name":"","customer_code":"","ship_to":"","bill_to":"",
                "terms":"","currency":"USD","lines":[],
                "warnings":["⚠️ pdfplumber no instalado. Corre: pip install pdfplumber"]}

    from collections import defaultdict, OrderedDict

    all_words = []
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page_num, page in enumerate(pdf.pages):
                for w in page.extract_words():
                    all_words.append({
                        'page': page_num, 'x0': w['x0'],
                        'top': w['top'], 'text': w['text']
                    })
    except Exception as e:
        return {"po_number":"","po_date":"","ship_date":"","cancel_date":"",
                "customer_name":"","customer_code":"","ship_to":"","bill_to":"",
                "terms":"","currency":"USD","lines":[],
                "warnings":[f"⚠️ No se pudo leer el PDF: {e}"]}

    if not all_words:
        return {"po_number":"","po_date":"","ship_date":"","cancel_date":"",
                "customer_name":"","customer_code":"","ship_to":"","bill_to":"",
                "terms":"","currency":"USD","lines":[],
                "warnings":["⚠️ El PDF no tiene texto extraíble (puede ser escaneado)."]}

    # Group words into lines by page+top (within 2px)
    lines_dict = defaultdict(list)
    for w in all_words:
        key = (w['page'], round(w['top']/2)*2)
        lines_dict[key].append(w)
    sorted_keys = sorted(lines_dict.keys())

    # Build text lines
    text_lines = []
    for key in sorted_keys:
        ws = sorted(lines_dict[key], key=lambda w: w['x0'])
        text_lines.append(' '.join(w['text'] for w in ws))

    # ── Header extraction ──────────────────────────────────────
    PDF_STOPS = [
        'Vendor:','Ship To:','Bill To:','PO Number:','PO Class:',
        'Version:','Last Submitted','Promised','Currency:','Total Qty:',
        'Total Cost:','DamageAllowance:','Buyer:','Phone:','Ship Via:',
        'Terms:','Cancel Date:','Ship Date:','Attn:','Special Instructions:'
    ]
    def find_pdf_val(lines, keywords):
        for line in lines:
            up = line.upper()
            for kw in keywords:
                idx = up.find(kw.upper())
                if idx != -1:
                    after = line[idx+len(kw):].strip().lstrip(':').strip()
                    # Stop at next known label
                    for stop in PDF_STOPS:
                        si = after.upper().find(stop.upper())
                        if si > 0:
                            after = after[:si].strip()
                    # Stop at 3+ spaces
                    after = re.split(r'\s{3,}', after)[0].strip()
                    if after and len(after) > 0:
                        return after
        return ''

    result = {
        "po_number":     find_pdf_val(text_lines, [kw for kw in GLOSSARY["po_number"]]),
        "po_date":       find_pdf_val(text_lines, [kw for kw in GLOSSARY["po_date"]]),
        "ship_date":     find_pdf_val(text_lines, [kw for kw in GLOSSARY["ship_date"]]),
        "cancel_date":   find_pdf_val(text_lines, [kw for kw in GLOSSARY["cancel_date"]]),
        "customer_name": find_pdf_val(text_lines, ["Bill To:", "Sold To:", "Customer:"]),
        "customer_code": "",
        "ship_to":       find_pdf_val(text_lines, ["Ship To:", "Delivery Address:"]),
        "bill_to":       find_pdf_val(text_lines, ["Bill To:"]),
        "buyer":         find_pdf_val(text_lines, ["Buyer:"]),
        "terms":         find_pdf_val(text_lines, ["Terms:"]),
        "currency":      find_pdf_val(text_lines, ["Currency:"]) or "USD",
        "lines":         [],
        "warnings":      [],
        "source":        "pdf",
    }

    # ── Detect PDF format ────────────────────────────────────────
    # Format A (Fanatics): style code appears directly in table as column
    # Format B (Delaware North): style in "Code: BBC1517454-WBK" lines below item
    # Format C (Dallas Cowboys): style "FDC1410291", size+color "2XL / WHT" in same line

    style_col_pattern = re.compile(r'^[A-Z]{3,6}\d{6,}-[A-Z]{2,4}$')
    code_line_pattern = re.compile(r'Code:\s+([A-Z]{3,6}\d{6,}-[A-Z]{2,4})')
    style_nohyp_pat   = re.compile(r'\b([A-Z]{3,6}\d{6,})\b')
    size_color_pat    = re.compile(r'\b(XXL|2XL|3XL|4XL|5XL|XS|XL|S|M|L)\s+/\s+([A-Z]+)\b')
    SIZE_RE = re.compile(r'^(XS|XXS|XXL|2XL|3XL|4XL|5XL|S|M|L|XL|OS|OSFA)$')

    has_code_lines = any(code_line_pattern.search(line) for line in text_lines)
    has_col_styles = any(style_col_pattern.match(w['text']) for w in all_words)
    has_size_color = any(size_color_pat.search(line) for line in text_lines)
    # Format D (Caesars): style+desc+size+$cost+qty+UPC all in one line
    caesars_pat = re.compile(r'[A-Z]{3,6}\d{6,}-[A-Z]{2,4}.+\$\d+\.\d{2}.+\d{12,}')
    has_caesars = any(caesars_pat.search(line) for line in text_lines)
    # Format E (Shiekh LAB): "P.O. #" header, style repeated twice, tallas en siguiente línea
    has_shiekh_lab = any(re.match(r'P\.O\. #\s+\S+', line) for line in text_lines)
    # Format F (Follett): "P/O Number -", style has "/" like "70102/CSNUCRD"
    has_follett = any('P/O Number' in line or 'P/O Location' in line for line in text_lines)
    # Format G (Blue Jays / PLU-VLU): columnas PLU VLU Product Description Colour Name Size Qty Unit Cost
    has_plu_vlu = any(re.search(r'\bPLU\b', line) and re.search(r'\bVLU\b', line) for line in text_lines)

    product_lines = []

    if has_code_lines:
        # ── Format B: Delaware North / "Code:" style ──────────
        # Also update header: PO# from "#NNNN" line
        # Reset header — Delaware North has its own format, override generic parsing
        result["po_number"] = ""; result["ship_date"] = ""; result["cancel_date"] = ""
        for line in text_lines:
            # PO# from "#NNNN" line
            m = re.match(r'^#(\d+)$', line.strip())
            if m and not result["po_number"]:
                result["po_number"] = m.group(1)
            # Dates from "Ship Date  Cancel Date  Buyer" header row
            if 'Ship Date' in line and 'Cancel Date' in line and 'Buyer' in line:
                idx = text_lines.index(line)
                if idx+1 < len(text_lines):
                    parts = text_lines[idx+1].split()
                    if len(parts) >= 1 and re.match(r'\d+/\d+/\d+', parts[0]):
                        result["ship_date"]   = parts[0]
                        if len(parts) >= 2: result["cancel_date"] = parts[1]
                        if len(parts) >= 3: result["buyer"]       = ' '.join(parts[2:])
            # Customer from "Vendor  Ship To  Bill To" header — use X positions
            if 'Bill To' in line and 'Ship To' in line and 'Vendor' in line:
                idx = text_lines.index(line)
                if idx+1 < len(text_lines):
                    next_key = sorted_keys[idx+1]
                    next_words = sorted(lines_dict[next_key], key=lambda w: w['x0'])
                    # Bill To is the rightmost column (x > 280)
                    bill_to_words = [w['text'] for w in next_words if w['x0'] > 280]
                    if bill_to_words:
                        result["customer_name"] = ' '.join(bill_to_words).strip()
                    ship_to_words = [w['text'] for w in next_words if 140 < w['x0'] <= 280]
                    if ship_to_words:
                        result["ship_to"] = ' '.join(ship_to_words).strip()

        price_re = re.compile(r'\$(\d+\.\d{2})')
        for i, line in enumerate(text_lines):
            code_match = code_line_pattern.search(line) if i > 0 else None
            if code_match:
                style    = code_match.group(1)
                prev     = text_lines[i-1]
                qty_m    = re.match(r'^(\d+)\s+', prev)
                qty      = int(qty_m.group(1)) if qty_m else 0
                # Description: text between qty and size (remove qty, size, prices)
                desc = prev
                if qty_m: desc = desc[qty_m.end():]
                dollar_i = desc.find('$')
                if dollar_i > 0: desc = desc[:dollar_i]
                # Remove trailing size (last word like S, M, XL, XXL)
                desc = re.sub(r'\s+(XXL|2XL|3XL|4XL|5XL|XS|XL|S|M|L)\s*$', '', desc).strip()
                # Take LAST size before $ sign (avoids description words like SS, BLK)
                dollar_i = prev.find('$')
                before_dollar = prev[:dollar_i].strip() if dollar_i > 0 else prev
                size_all = re.findall(r'\b(XXL|2XL|3XL|4XL|5XL|XS|S|M|L|XL|OS)\b', before_dollar)
                size     = size_all[-1].replace('XXL','2XL') if size_all else ''
                prices   = price_re.findall(prev)
                cost     = float(prices[0]) if prices else 0
                msrp     = float(prices[1]) if len(prices) > 1 else 0
                if qty > 0:
                    product_lines.append({'style':style,'size':size,'qty':qty,'cost':cost,'msrp':msrp,'desc':desc})

    elif has_caesars:
        # ── Format D: Caesars / style+desc+size+$cost+qty+UPC in one line ──
        result["po_number"] = ""; result["ship_date"] = ""; result["cancel_date"] = ""
        for line in text_lines:
            m = re.search(r'Purchase Order:\s+(\S+)', line)
            if m and not result["po_number"]: result["po_number"] = m.group(1)
            m = re.search(r'Start Ship:\s+(\S+)', line)
            if m and not result["ship_date"]: result["ship_date"] = m.group(1)
            m = re.search(r'Cancel Date:\s+(\S+)', line)
            if m and not result["cancel_date"]: result["cancel_date"] = m.group(1)
            m = re.search(r'Bill to:\s+(.+)', line)
            if m and not result["customer_name"]:
                result["customer_name"] = m.group(1).strip()
            # Delivery address
            if 'Delivery Address:' in line and not result["ship_to"]:
                m2 = re.search(r'Delivery Address:\s+(.+)', line)
                if m2:
                    result["ship_to"] = m2.group(1).strip()

        caesars_size_pat = re.compile(r'[A-Z]*(MD|LG|2XL|3XL|XXL|2X|3X|XL|XS|SM|S|M|L)\s+\$')
        caesars_price_pat = re.compile(r'\$([\d,]+\.\d{2})')
        caesars_qty_pat   = re.compile(r'\b(\d{1,3})\s+\d{12,}')
        caesars_sz_norm   = {"MD":"M","LG":"L","SM":"S","XL":"XL","2X":"2XL","3X":"3XL",
                             "XXL":"2XL","2XL":"2XL","3XL":"3XL","XS":"XS","S":"S","M":"M","L":"L"}

        for line in text_lines:
            sm = style_col_pattern.search(line) if False else None
            # Use style with hyphen pattern
            st_m = re.search(r'\b([A-Z]{3,6}\d{6,}-[A-Z]{2,4})\b', line)
            sz_m = caesars_size_pat.search(line)
            if not st_m or not sz_m: continue
            prices = caesars_price_pat.findall(line)
            qty_m  = caesars_qty_pat.search(line)
            style  = st_m.group(1)
            raw_sz = sz_m.group(1)
            size   = caesars_sz_norm.get(raw_sz, raw_sz)
            cost   = float(prices[0].replace(',','')) if prices else 0
            msrp   = float(prices[1].replace(',','')) if len(prices) > 1 else 0
            qty    = int(qty_m.group(1)) if qty_m else 0
            # Description: between style and size
            between = line[st_m.end():sz_m.start()].strip()
            words   = between.split()
            desc    = ' '.join(words[:-2]) if len(words) > 2 else ' '.join(words[:-1])
            if qty > 0:
                product_lines.append({'style':style,'size':size,'qty':qty,
                                      'cost':cost,'msrp':msrp,'desc':desc})


    elif has_col_styles:
        # ── Format A: Fanatics / style as column ──────────────
        col_x = {"style": 399, "size": 620, "qty": 691, "msrp": 724, "cost": 764}
        for key in sorted_keys:
            ws = sorted(lines_dict[key], key=lambda w: w['x0'])
            texts = [w['text'].upper() for w in ws]
            line  = ' '.join(texts)
            if 'VENDOR STYLE' in line and ('SIZE' in line or 'QTY' in line):
                for w in ws:
                    t = w['text'].upper()
                    if matches_any(t, GLOSSARY["col_style"]):   col_x["style"] = w['x0']
                    if matches_any(t, ["SIZE"]):                col_x["size"]  = w['x0']
                    if matches_any(t, GLOSSARY["col_qty"]):     col_x["qty"]   = w['x0']
                    if matches_any(t, GLOSSARY["col_msrp"]):    col_x["msrp"]  = w['x0']
                    if matches_any(t, GLOSSARY["col_cost"]):    col_x["cost"]  = w['x0']
                break
        COL_TOL = 25
        for key in sorted_keys:
            ws = sorted(lines_dict[key], key=lambda w: w['x0'])
            style_w = next((w for w in ws if style_col_pattern.match(w['text'])), None)
            if not style_w: continue
            style = style_w['text']; size = ''; qty = 0; msrp = 0.0; cost = 0.0
            for w in ws:
                t = w['text']; x = w['x0']
                if abs(x-col_x["size"]) <= COL_TOL and SIZE_RE.match(t.upper()):
                    size = t.upper().replace('XXL','2XL')
                elif abs(x-col_x["qty"]) <= COL_TOL and re.match(r'^\d+$', t):
                    qty = int(t)
                elif abs(x-col_x["msrp"]) <= COL_TOL and re.match(r'^\d+\.\d+$', t):
                    msrp = float(t)
                elif abs(x-col_x["cost"]) <= COL_TOL and re.match(r'^\d+\.\d+$', t):
                    cost = float(t)
            if qty > 0:
                product_lines.append({'style':style,'size':size,'qty':qty,'msrp':msrp,'cost':cost,'desc':''})

    elif has_size_color:
        # ── Format C: Dallas Cowboys / "Style Id" + "Size / Color" in same line ──
        result["po_number"] = ""; result["ship_date"] = ""; result["cancel_date"] = ""
        for i, line in enumerate(text_lines):
            m = re.match(r'Purchase Order\s+(PO\w+)', line)
            if m and not result["po_number"]: result["po_number"] = m.group(1)
            m = re.match(r'PO Date\s+(\S+)', line)
            if m: result["po_date"] = m.group(1)
            m = re.match(r'Cancel Date\s+(\S+)', line)
            if m: result["cancel_date"] = m.group(1)
            if line.strip() == 'Delivery Date' and i > 0:
                prev = text_lines[i-1].strip()
                if re.match(r'\d+/\d+/\d+', prev): result["ship_date"] = prev
            if 'Vendor' in line and 'Delivery Address' in line and i+1 < len(text_lines):
                parts = re.split(r'\s{2,}', text_lines[i+1])
                if len(parts) >= 2: result["customer_name"] = parts[-1].strip()

        # Customer — look for "Delivery Address" label then find value after it
        if not result["customer_name"]:
            for i, line in enumerate(text_lines):
                if line.strip() == 'Delivery Address' and i+1 < len(text_lines):
                    # Next line is the delivery address value
                    val = text_lines[i+1].strip()
                    if val and not re.match(r'[\d/]+', val) and len(val) > 3:
                        result["customer_name"] = val
                    break
                # Also: "Vendor Delivery Address" on same line — value in NEXT line right part
                if 'Vendor' in line and 'Delivery Address' in line and i+1 < len(text_lines):
                    # Scan next lines for the delivery store name
                    for nl in text_lines[i+1:i+5]:
                        # Delivery address lines come AFTER vendor address
                        # Look for a line that has a store name format (not PO info)
                        if re.search(r'\bPro Shop\b|\bArena\b|\bWarehouse\b|\bStore\b|\bGalleria\b', nl, re.I):
                            result["customer_name"] = nl.strip()
                            break
                    break

        qty_re = re.compile(r'(\d+)\.00\s+EA')
        for line in text_lines:
            style_m = style_nohyp_pat.search(line)
            sc_m    = size_color_pat.search(line)
            if not style_m or not sc_m: continue
            qty_m  = qty_re.search(line)
            prices = re.findall(r'\d+\.\d{2}', line)
            style  = style_m.group(1)
            size   = sc_m.group(1).replace('XXL','2XL')
            color  = sc_m.group(2)
            qty    = int(qty_m.group(1)) if qty_m else 0
            cost   = float(prices[1]) if len(prices) > 1 else (float(prices[0]) if prices else 0)
            desc_m = re.match(r'^\d+\s+\d+\s+([A-Z][A-Z\s/]+?)\s+FDC', line)
            desc   = desc_m.group(1).strip() if desc_m else ''
            if qty > 0:
                product_lines.append({'style':f"{style}-{color}",'size':size,'qty':qty,'cost':cost,'msrp':0,'desc':desc})


    elif has_shiekh_lab:
        # ── Format E: Shiekh LAB / "P.O. #" header, style repeated twice ──
        result["po_number"] = ""; result["ship_date"] = ""; result["cancel_date"] = ""
        # PO# from "P.O. # SHK-LABSACC-JULY26"
        for line in text_lines:
            m = re.match(r'P\.O\. #\s+(\S+)', line)
            if m and not result["po_number"]:
                result["po_number"] = m.group(1)
            m = re.search(r'Ship Date:\s+(\S+)', line)
            if m and not result["ship_date"]: result["ship_date"] = m.group(1)
            m = re.search(r'Cancel Date:\s+(\S+)', line)
            if m and not result["cancel_date"]: result["cancel_date"] = m.group(1)
            # Customer from "TO: LAB LAB co OFF WHITE SHIP TO: Shiekh Shoes"
            m = re.search(r'SHIP TO:\s+(.+)', line)
            if m and not result["customer_name"]:
                result["customer_name"] = m.group(1).strip()

        shiekh_style_pat = re.compile(r'^([A-Z]{2,6}\d{3,}[-]\w+)\s')
        size_norm_sh = {"1SIZE":"OS","OSFA":"OS","M":"M","L":"L","S":"S","XL":"XL",
                        "XXL":"2XL","2XL":"2XL","3XL":"3XL","XS":"XS","SM":"S","MD":"M","LG":"L"}
        i = 0
        while i < len(text_lines):
            line = text_lines[i]
            sm = shiekh_style_pat.match(line)
            if sm:
                style = sm.group(1)
                # Find second occurrence for cost/qty
                second_idx = line.find(style, sm.end())
                if second_idx > 0:
                    after = line[second_idx + len(style):].strip()
                    nums = re.findall(r'[\d,]+\.?\d*', after)
                    nums_f = [float(n.replace(',','')) for n in nums]
                    cost = nums_f[0] if nums_f else 0
                    qty_total = int(nums_f[1]) if len(nums_f) > 1 else 0
                    msrp = nums_f[-1] if len(nums_f) > 3 else 0
                else:
                    cost = 0; qty_total = 0; msrp = 0
                # Get description
                desc_start = sm.end()
                desc_end = line.find(style, desc_start) if second_idx > 0 else len(line)
                desc = line[desc_start:desc_end].strip()
                # Next lines: sizes + quantities
                sizes = {}
                if i+2 < len(text_lines):
                    size_line = text_lines[i+1].strip()
                    qty_line  = text_lines[i+2].strip()
                    # Check if they look like sizes
                    sz_words = size_line.split()
                    qt_words = qty_line.split()
                    if all(w.upper() in size_norm_sh or w == "1SIZE" for w in sz_words):
                        if size_line.upper() in ("1SIZE","OSFA","ONE SIZE"):
                            try: sizes["OS"] = int(qty_line)
                            except: sizes["OS"] = qty_total
                        else:
                            for sz, qv in zip(sz_words, qt_words):
                                norm = size_norm_sh.get(sz.upper(), sz.upper())
                                try: sizes[norm] = int(qv)
                                except: pass
                        i += 3
                        total_units = sum(sizes.values())
                    else:
                        sizes["OS"] = qty_total
                        total_units = qty_total
                        i += 1
                else:
                    sizes["OS"] = qty_total
                    total_units = qty_total
                    i += 1
                if total_units > 0:
                    product_lines.append({
                        'style': style, 'size': '', 'qty': total_units,
                        'cost': cost, 'msrp': msrp, 'desc': desc,
                        'sizes_map': sizes
                    })
            else:
                i += 1

    elif has_plu_vlu:
        # ── Format G: Blue Jays / PLU-VLU — una fila por talla ──
        # VLU puede estar partido en dos líneas: "LTJ1315907-" + "ERB" en la siguiente
        # Header: "PO #: 3328", fechas en fila "Order Date Ship Date Arrival Date Cancel Date"
        result["po_number"] = ""; result["ship_date"] = ""; result["cancel_date"] = ""
        result["customer_name"] = ""

        # Customer: línea justo después de "Purchase Order"
        for i, line in enumerate(text_lines):
            if line.strip() == 'Purchase Order' and i+1 < len(text_lines):
                result["customer_name"] = text_lines[i+1].strip()
                break

        for i, line in enumerate(text_lines):
            m = re.search(r'PO\s*#[:\s]+(\w+)', line)
            if m and not result["po_number"]:
                result["po_number"] = m.group(1)
            m = re.search(r'Buyer[:\s]+(.+)', line)
            if m and not result.get("buyer"):
                result["buyer"] = m.group(1).strip().split('  ')[0]
            if re.search(r'Order Date', line) and re.search(r'Ship Date', line) and re.search(r'Cancel Date', line):
                if i+1 < len(text_lines):
                    date_vals = re.findall(r'\d+/\d+/\d+', text_lines[i+1])
                    if len(date_vals) >= 2: result["ship_date"]   = date_vals[1]
                    if len(date_vals) >= 4: result["cancel_date"] = date_vals[3]

        # Find header row index
        header_idx = None
        for i, line in enumerate(text_lines):
            if re.search(r'\bPLU\b', line) and re.search(r'\bVLU\b', line):
                header_idx = i
                break

        if header_idx is not None:
            size_norm_g = {"S":"S","M":"M","L":"L","XL":"XL","2XL":"2XL","3XL":"3XL",
                           "4XL":"4XL","5XL":"5XL","XS":"XS","XXS":"XXS","OS":"OS","OSFA":"OS"}
            i = header_idx + 1
            while i < len(text_lines):
                line = text_lines[i]
                # Stop at totals/legal
                if re.match(r'^\s*(Total:|CHANGES|AT ANY|Printed:)', line):
                    break
                # Data row starts with 6-digit PLU
                m_plu = re.match(r'^(\d{6})\s+', line)
                if not m_plu:
                    i += 1; continue

                # VLU: may be "LTJ1315907-" (trailing hyphen = split across lines)
                vlu = ''
                after_plu = line[m_plu.end():]
                vlu_m = re.match(r'([A-Z0-9]+-[A-Z]*)', after_plu)
                if vlu_m:
                    vlu_raw = vlu_m.group(1)
                    if vlu_raw.endswith('-'):
                        # Suffix is on the next line
                        if i+1 < len(text_lines):
                            suffix_m = re.match(r'^([A-Z]+)\b', text_lines[i+1].strip())
                            if suffix_m:
                                vlu = vlu_raw + suffix_m.group(1)
                    else:
                        vlu = vlu_raw

                if not vlu:
                    i += 1; continue

                # Numbers at end: Qty, Item Retail, Unit Cost, Ext.Cost
                nums = re.findall(r'[\d,]+\.?\d*', line)
                # Remove PLU (first 6-digit number)
                nums = [n for n in nums if not re.match(r'^\d{6}$', n)]
                nums_f = [float(n.replace(',','')) for n in nums]
                qty         = int(nums_f[-4]) if len(nums_f) >= 4 else 0
                item_retail = nums_f[-3]      if len(nums_f) >= 3 else 0
                unit_cost   = nums_f[-2]      if len(nums_f) >= 2 else 0

                # Size: find size token in the line
                size = "OS"
                for p in reversed(line.split()):
                    if p.upper() in size_norm_g:
                        size = size_norm_g[p.upper()]; break

                # Description: everything between VLU and the size token
                vlu_stub = vlu_raw  # what appeared in this line (may have trailing -)
                vlu_pos = line.find(vlu_stub)
                desc_raw = line[vlu_pos + len(vlu_stub):].strip()
                # Strip trailing numbers and size
                desc_raw = re.sub(r'\s+\S+\s+[\d,]+\.?\d*\s+[\d,]+\.?\d*\s+[\d,]+\.?\d*\s+[\d,]+\.?\d*\s*$', '', desc_raw).strip()

                if qty > 0:
                    product_lines.append({
                        'style': vlu, 'size': size, 'qty': qty,
                        'cost': unit_cost, 'msrp': item_retail, 'desc': desc_raw
                    })
                i += 1

    elif has_follett:
        # ── Format F: Follett / "P/O Number -", style with "/", size "Color / Size" ──
        result["po_number"] = ""; result["ship_date"] = ""; result["cancel_date"] = ""
        result["customer_name"] = ""  # reset — generic block picks up wrong value
        for line in text_lines:
            m = re.search(r'P/O Number\s*-\s*(\S+)', line)
            if m and not result["po_number"]: result["po_number"] = m.group(1)
            m = re.search(r'Ship Date\s*-\s*(\S+)', line)
            if m and not result["ship_date"]: result["ship_date"] = m.group(1)
            m = re.search(r'Cancel Date\s*-\s*(\S+)', line)
            if m and not result["cancel_date"]: result["cancel_date"] = m.group(1)
            if 'Ship To:' in line and not result["customer_name"]:
                idx = text_lines.index(line)
                if idx+1 < len(text_lines):
                    result["ship_to"] = text_lines[idx+1].strip()
                    result["customer_name"] = result["ship_to"]

        follett_style_pat = re.compile(r'^\d+\s+(\S+/\S+)\s+')
        follett_size_map = {
            "XSMALL":"XS","XXSMALL":"XXS","SMALL":"S","MEDIUM":"M","LARGE":"L",
            "XLARGE":"XL","XXLARGE":"2XL","2XLARGE":"2XL","XXXLARGE":"3XL","3XLARGE":"3XL",
            "XS":"XS","S":"S","M":"M","L":"L","XL":"XL","XXL":"2XL","2XL":"2XL",
        }
        for i, line in enumerate(text_lines):
            fm = follett_style_pat.match(line)
            if not fm: continue
            style = fm.group(1)
            nums = re.findall(r'\d+\.?\d*', line)
            nums_f = [float(n) for n in nums]
            qty = int(nums_f[0]) if nums_f else 0
            cost = next((n for n in nums_f[1:] if 0 < n < 10000 and n != int(n)), 0)
            msrp = next((n for n in reversed(nums_f) if n != cost and n > 0), 0)
            # Size from line i+2 "Color / Size" (line i+1 is UPC)
            size = "OS"
            for offset in range(1, 8):
                if i+offset >= len(text_lines): break
                nxt = text_lines[i+offset]
                # Skip UPC, /Cont., page headers, PO info lines
                if any(skip in nxt for skip in ['UPC', '/Cont.', 'POM345', 'Sub Total', 'P/O', 'Purchase Order', '*EMAIL', 'Q t y', 'U n i t']):
                    continue
                # Only match "Word / Word" pattern (Color / Size)
                if re.match(r'^[A-Za-z\s]+ / [A-Za-z]+', nxt.strip()):
                    sz_raw = nxt.split('/')[-1].strip().upper().replace(' ','')
                    mapped = follett_size_map.get(sz_raw, sz_raw)
                    if mapped and mapped != 'CONT.':
                        size = mapped
                        break
            if qty > 0:
                product_lines.append({
                    'style': style, 'size': size, 'qty': qty,
                    'cost': cost, 'msrp': msrp, 'desc': ''
                })

    # Group by style+color
    grouped = OrderedDict()
    for pl in product_lines:
        k = pl['style']
        if k not in grouped:
            grouped[k] = {
                'style': k, 'sizes': {}, 'cost': pl['cost'],
                'msrp': pl['msrp'], 'total': 0, 'desc': pl.get('desc','')
            }
        # Shiekh LAB has pre-built sizes_map
        if pl.get('sizes_map'):
            for sz, qty in pl['sizes_map'].items():
                grouped[k]['sizes'][sz] = grouped[k]['sizes'].get(sz, 0) + qty
            grouped[k]['total'] += pl['qty']
        else:
            sz = pl['size'] or 'OS'
            grouped[k]['sizes'][sz] = grouped[k]['sizes'].get(sz, 0) + pl['qty']
            grouped[k]['total'] += pl['qty']
        if not grouped[k]['cost'] and pl['cost']:
            grouped[k]['cost'] = pl['cost']
        if not grouped[k]['desc'] and pl.get('desc'):
            grouped[k]['desc'] = pl['desc']

    # Convert to standard line format
    for data in grouped.values():
        stock, cc, cn = parse_style_color(data['style'])
        sizes = {s: 0 for s in SIZE_ORDER}
        for sz, qty in data['sizes'].items():
            norm = normalize_size(sz)
            if norm in sizes:
                sizes[norm] = qty
            else:
                sizes[sz] = qty
        total_units = data['total']
        line_cost   = data['cost']
        result["lines"].append({
            'style_raw': data['style'], 'stock': stock,
            'color_code': cc, 'color_name': cn,
            'description': data.get('desc',''), 'cost': data['cost'], 'msrp': data['msrp'],
            'discount': 0, 'line_cost': line_cost,
            'sizes': sizes, 'total_units': total_units,
            'total_cost': line_cost * total_units,
            'total_retail': data['msrp'] * total_units,
        })

    if not result["lines"]:
        result["warnings"].append(
            "⚠️ No se detectaron líneas de producto. "
            "Verifica que el PDF tenga texto (no sea escaneado) o contacta soporte."
        )

    return result

# ══════════════════════════════════════════════════════════════
# PARSER PRINCIPAL
# ══════════════════════════════════════════════════════════════

def parse_po_excel(file_bytes):
    # Support both .xlsx and legacy .xls
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        all_sheet_rows = [(sn, [[cell.value for cell in row] for row in wb[sn].iter_rows()])
                          for sn in wb.sheetnames]
    except Exception:
        try:
            import xlrd
            wb_xls = xlrd.open_workbook(file_contents=file_bytes)
            ws_xls = wb_xls.sheet_by_index(0)
            from datetime import datetime as _dt
            rows = []
            for i in range(ws_xls.nrows):
                row = []
                for j in range(ws_xls.ncols):
                    cell = ws_xls.cell(i, j)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            dt_tuple = xlrd.xldate_as_tuple(cell.value, wb_xls.datemode)
                            row.append(_dt(*dt_tuple[:6]) if dt_tuple[0] > 0 else None)
                        except:
                            row.append(cell.value)
                    elif cell.ctype == xlrd.XL_CELL_NUMBER:
                        v = cell.value
                        row.append(int(v) if v == int(v) else v)
                    elif cell.ctype == xlrd.XL_CELL_EMPTY:
                        row.append(None)
                    else:
                        row.append(cell.value)
                rows.append(row)
        except Exception as e2:
            return {"po_number":"","po_date":"","ship_date":"","cancel_date":"",
                    "customer_name":"","customer_code":"","ship_to":"","bill_to":"",
                    "terms":"","currency":"USD","lines":[],
                    "warnings":[f"⚠️ No se pudo leer el archivo: {e2}"]}
        all_sheet_rows = [("Sheet1", rows)]  # xls: single sheet

    result = {
        "po_number":"","po_date":"","ship_date":"","cancel_date":"",
        "customer_name":"","customer_code":"","ship_to":"","bill_to":"",
        "terms":"","currency":"USD","lines":[],"warnings":[]
    }

    # ── 1. Header ────────────────────────────────────────────
    po_number = find_header_value(rows, GLOSSARY["po_number"])
    if not po_number:
        for row in rows[:5]:
            for cell in row:
                if cell and "PO" in str(cell).upper():
                    # Only match if cell STARTS with PO (avoid matching "PO" inside descriptions)
                    m = re.match(r'PO#?:?\s+([A-Z0-9][A-Z0-9\.\-]+)', str(cell).strip(), re.IGNORECASE)
                    if m:
                        candidate = m.group(1).strip()
                        if len(candidate) > 2 and candidate.upper() not in ("NAME","NUMBER","DATE","NO","HOODIE","TEE","JACKET"):
                            po_number = candidate
                            break
            if po_number:
                break
    result["po_number"]   = str(po_number).strip() if po_number else ""
    result["po_date"]     = fmtDate(find_header_value(rows, GLOSSARY["po_date"]))
    ship_raw = find_header_raw(rows, GLOSSARY["ship_date"])
    result["ship_date"] = fmtDate(ship_raw) if ship_raw else ""
    result["cancel_date"] = fmtDate(find_header_value(rows, GLOSSARY["cancel_date"]))
    result["terms"]       = find_header_value(rows, GLOSSARY["terms"])

    # Customer name — special BILL TO block first (handles both formats):
    # Format A: "Bill To:   FITTERS" — label+value in same cell (Fitters)
    # Format B: "BILL TO:" label, value on next row same col (Kings)
    customer_name = ""
    for i, row in enumerate(rows[:15]):
        for j, cell in enumerate(row):
            raw = str(cell or "").strip()
            c   = raw.upper()
            # Check if this cell starts with a BILL TO / SOLD TO variant
            is_bill_label = any(c.startswith(kw) for kw in [
                "BILL TO","BILL-TO","SOLD TO","BILLED TO"
            ])
            if not is_bill_label:
                continue
            # Format A: value after colon in same cell e.g. "Bill To:   FITTERS"
            if ":" in raw:
                after = raw.split(":",1)[1].strip()
                if after and len(after) > 2 and after.upper() not in _LABEL_WORDS:
                    customer_name = after
                    break
            # Format B2: value in next column same row (Guerreros, Borregos, Maxima template)
            # e.g. col0="Bill To:" col1="DEPORTES MARTI"
            if j+1 < len(row) and row[j+1] is not None:
                v = str(row[j+1]).strip()
                if v and len(v) > 2 and v.upper() not in _LABEL_WORDS:
                    customer_name = v
                    break
            # Format C: value on next row, same column (Kings)
            if i+1 < len(rows) and j < len(rows[i+1]):
                v = str(rows[i+1][j] or "").strip()
                if v and len(v) > 2 and v.upper() not in _LABEL_WORDS:
                    customer_name = v
                    break
        if customer_name:
            break
    # Fallback to generic glossary if not found — search full header (up to 60 rows)
    if not customer_name:
        customer_name = find_header_value(rows, GLOSSARY["customer_name"], max_rows=60)
    # Last resort: extract from title row containing "PURCHASE ORDER"
    if not customer_name:
        for row in rows[:3]:
            for cell in row:
                if cell and "PURCHASE ORDER" in str(cell).upper():
                    raw = str(cell).strip().lstrip("'")
                    # Remove "PURCHASE ORDER FORM" suffix
                    name = re.sub(r"PURCHASE ORDER.*", "", raw, flags=re.IGNORECASE).strip()
                    if name and len(name) > 3:
                        customer_name = name
                        break
            if customer_name:
                break
    result["customer_name"] = customer_name
    result["customer_code"] = find_header_value(rows, GLOSSARY["customer_code"])

    ship_to = find_header_value(rows, GLOSSARY["ship_to"], max_rows=60)
    if not ship_to:
        for i, row in enumerate(rows):  # search full file
            for j, cell in enumerate(row):
                c = cell_str(cell)
                if matches_any(c, ["SHIP TO:","SHIP TO","SHIPPING ADDRESS:","DELIVERY ADDRESS:"]):
                    # Value after colon in same cell
                    raw = str(cell).strip()
                    if ":" in raw:
                        after = raw.split(":",1)[1].strip()
                        if after and len(after) > 2 and after.upper() not in _LABEL_WORDS:
                            ship_to = after; break
                    # Next rows same column
                    parts = []
                    for k in range(i+1, min(i+5, len(rows))):
                        if j < len(rows[k]):
                            v = str(rows[k][j] or "").strip()
                            if v.upper() in ("DESCRIPTION","VENDOR STYLE NO.","STYLE","QTY","UPC","STYLE NUMBER"):
                                break
                            if v:
                                parts.append(v)
                    if parts:
                        ship_to = " | ".join(parts)
                    break
            if ship_to:
                break
    result["ship_to"] = ship_to
    result["bill_to"] = find_header_value(rows, GLOSSARY["bill_to"])

    # ── 2. Encontrar tabla de productos ──────────────────────
    data_header_row = None
    col_map = {}

    # ── USSF format detection ──────────────────────────────────
    is_ussf = False
    for test_row in rows[:6]:
        if len(test_row) > 6 and cell_str(test_row[6]) in ("DELIVER","CXL","PO"):
            is_ussf = True
            break

    # ── Shiekh format detection ─────────────────────────────────
    # Shiekh: SHIP in col 20, CANCEL in col 20, PO# in col 17, tallas en row 5 cols 8-16
    is_shiekh = False
    for test_row in rows[:5]:
        if len(test_row) > 20 and cell_str(test_row[20]) in ("SHIP","CANCEL"):
            is_shiekh = True
            break
    if is_shiekh:
        for row in rows[:6]:
            if len(row) > 21:
                if cell_str(row[20]) == "SHIP":
                    result["ship_date"] = fmtDate(row[21])
                elif cell_str(row[20]) == "CANCEL":
                    result["cancel_date"] = fmtDate(row[21])
                if cell_str(row[17]) == "PO#":
                    result["po_number"] = str(row[18] or "").strip()
        # Customer from row 1 col 3
        if len(rows) > 1 and len(rows[1]) > 3 and rows[1][3]:
            result["customer_name"] = str(rows[1][3]).strip()
        if len(rows) > 2 and len(rows[2]) > 5 and rows[2][5]:
            result["ship_to"] = str(rows[2][5]).strip()
    if is_ussf:
        # Extract USSF header fields
        for row in rows[:6]:
            if len(row) > 7:
                if cell_str(row[6]) == "PO":
                    result["po_number"] = str(row[7] or "").strip()
                elif cell_str(row[6]) == "DELIVER":
                    result["ship_date"] = fmtDate(row[7])
                elif cell_str(row[6]) == "CXL":
                    result["cancel_date"] = fmtDate(row[7])
        # Customer from row 1 col 0
        if len(rows) > 1 and rows[1][0]:
            result["customer_name"] = str(rows[1][0]).strip()
        # Ship to same as customer
        result["ship_to"] = result["customer_name"]

    for i, row in enumerate(rows):
        row_str = " | ".join(cell_str(c) for c in row)
        has_style = matches_any(row_str, GLOSSARY["col_style"])
        has_qty   = (matches_any(row_str, GLOSSARY["col_qty"]) or
                     matches_any(row_str, GLOSSARY["col_size_break"]) or
                     any(matches_any(row_str, GLOSSARY[k]) for k in SIZE_KEY_MAP))
        # USSF: detect header by ITEM # in col 0 and QTY. in col 8
        if is_ussf and not has_qty:
            if cell_str(row[0]) in ("ITEM #","ITEM#") if row[0] else False:
                has_style = True
                has_qty   = True
        if has_style and has_qty:
            data_header_row = i
            col_map["style"]        = detect_col(row, GLOSSARY["col_style"])
            col_map["desc"]         = detect_col(row, GLOSSARY["col_desc"])
            col_map["color"]        = detect_col(row, GLOSSARY["col_color"])
            col_map["size_break"]   = detect_col(row, GLOSSARY["col_size_break"])
            col_map["size"]         = detect_col(row, ["SIZE"])
            col_map["qty"]          = detect_col(row, GLOSSARY["col_qty"])
            # Detect cost columns — prefer net (COST W/DISC) over gross (COST)
            net_cost_keywords  = ["COST W/ DISC","COST W/DISC","COST W/DISC.","COST WITH DISCOUNT","NET COST","NET UNIT COST","DISCOUNTED PRICE","WHOLESALE NET","LINE COST","SALE PRICE"]
            gross_cost_keywords = ["UNIT COST","UNIT PRICE","PRICE PER UNIT","WHOLESALE","COST","WS","PRICE","COSTO UNITARIO","MAYOREO NETO"]
            col_net  = detect_col(row, net_cost_keywords)
            col_gross = detect_col(row, gross_cost_keywords)
            if col_net is not None:
                col_map["cost"] = col_net
                col_map["cost_is_net"] = True
            else:
                col_map["cost"] = col_gross
                col_map["cost_is_net"] = False
            col_map["msrp"]         = detect_col(row, GLOSSARY["col_msrp"])
            col_map["discount"]     = detect_col(row, GLOSSARY["col_discount"], exact_only=True)
            col_map["total_cost"]   = detect_col(row, GLOSSARY["col_total_cost"])
            col_map["total_retail"] = detect_col(row, GLOSSARY["col_total_retail"])
            # Detect size columns in this row
            for sz_key in SIZE_KEY_MAP:
                col_map[sz_key] = detect_col(row, GLOSSARY[sz_key], exact_only=True)
            # Also check NEXT row for size columns (two-row header like Fitters)
            # Only if next row has NO style code (i.e. it's a header row, not a data row)
            if i+1 < len(rows):
                next_row = rows[i+1]
                next_row_str = " | ".join(cell_str(c) for c in next_row)
                # Check if next row is a pure header row (no style codes)
                style_col = col_map.get("style")
                next_has_style_data = (
                    style_col is not None and
                    style_col < len(next_row) and
                    next_row[style_col] is not None and
                    re.match(r'^[A-Za-z]{2,}\d+', str(next_row[style_col]).strip())
                )
                if not next_has_style_data and any(matches_any(next_row_str, GLOSSARY[k]) for k in SIZE_KEY_MAP):
                    for sz_key in SIZE_KEY_MAP:
                        if col_map.get(sz_key) is None:
                            col_map[sz_key] = detect_col(next_row, GLOSSARY[sz_key], exact_only=True)
                    # Skip this second header row when reading data
                    data_header_row = i+1
            # USSF override: SIZE in col 6, QTY in col 7
            if is_ussf:
                col_map["size"] = 6
                col_map["qty"]  = 7
                col_map["cost"] = 9
                col_map["msrp"] = 10
                col_map["cost_is_net"] = False
            break

    if data_header_row is None:
        result["warnings"].append("⚠️ No se encontró la tabla de productos.")
        return result

    has_matrix     = any(col_map.get(k) is not None for k in SIZE_KEY_MAP)
    has_size_col   = col_map.get("size") is not None
    has_size_break = col_map.get("size_break") is not None

    # JD Sports format: ship/cancel dates in product table columns
    # Detect ship_date column in header row
    if data_header_row is not None:
        header_row_data = rows[data_header_row]
        ship_col   = detect_col(header_row_data, GLOSSARY["ship_date"])
        cancel_col = detect_col(header_row_data, GLOSSARY["cancel_date"])
        # Read from first data row
        if ship_col is not None and not result["ship_date"]:
            for dr in rows[data_header_row+1:data_header_row+3]:
                if ship_col < len(dr) and dr[ship_col]:
                    result["ship_date"] = fmtDate(dr[ship_col])
                    break
        if cancel_col is not None and not result["cancel_date"]:
            for dr in rows[data_header_row+1:data_header_row+3]:
                if cancel_col < len(dr) and dr[cancel_col]:
                    result["cancel_date"] = fmtDate(dr[cancel_col])
                    break

    # ── 3. Leer filas de productos ───────────────────────────
    lines = []
    last_valid_style = ""

    for row in rows[data_header_row + 1:]:
        row_text = " ".join(cell_str(c) for c in row if c)

        # Detectar ship_date embebida en fila (ej. "Ship Date: 10/1/2026")
        if matches_any(row_text, GLOSSARY["ship_date"]) and not result["ship_date"]:
            for c in row:
                if c and matches_any(cell_str(c), GLOSSARY["ship_date"]):
                    raw = str(c).strip()
                    if ":" in raw:
                        val = raw.split(":",1)[1].strip()
                        if val:
                            result["ship_date"] = fmtDate(val)
                            break
            continue

        # Obtener style
        style_raw = ""
        if col_map.get("style") is not None and col_map["style"] < len(row):
            style_raw = str(row[col_map["style"]] or "").strip()

        if style_raw:
            if matches_any(style_raw, ["TOTAL","SUBTOTAL","GRAND TOTAL","SUB-TOTAL","ABOVE TOTAL"]):
                continue
            if re.match(r'^[A-Za-z]{2,}\d+', style_raw):
                last_valid_style = style_raw
            else:
                continue
        else:
            if last_valid_style and (has_size_col or has_matrix):
                style_raw = last_valid_style
            else:
                continue

        stock, color_code, color_name = parse_style_color(style_raw)

        desc = ""
        if col_map.get("desc") is not None and col_map["desc"] < len(row):
            desc = str(row[col_map["desc"]] or "").strip()

        if col_map.get("color") is not None and col_map["color"] < len(row):
            cv = str(row[col_map["color"]] or "").strip()
            if cv:
                color_name = cv

        cost = 0.0
        if col_map.get("cost") is not None and col_map["cost"] < len(row):
            try: cost = float(row[col_map["cost"]] or 0)
            except: pass

        msrp = 0.0
        if col_map.get("msrp") is not None and col_map["msrp"] < len(row):
            try: msrp = float(row[col_map["msrp"]] or 0)
            except: pass

        discount = 0.0
        disc_col = col_map.get("discount")
        if disc_col is None:
            for idx in range(len(row)-1, -1, -1):
                v = row[idx]
                if v is not None:
                    try:
                        d = float(v)
                        if 0 < d <= 1:
                            disc_col = idx
                    except: pass
                    break
        if disc_col is not None and disc_col < len(row):
            try:
                d = float(row[disc_col] or 0)
                discount = d if d <= 1 else d / 100
            except: pass

        # ── Tallas ───────────────────────────────────────────
        sizes = {s: 0 for s in SIZE_ORDER}
        total_units = 0
        size_val = ""
        qty_val  = 0

        if has_size_break and col_map.get("size_break") is not None and col_map["size_break"] < len(row):
            sb = row[col_map["size_break"]]
            if sb:
                sizes = parse_size_break_compressed(sb)
                total_units = sum(sizes.values())

        elif has_matrix:
            for sz_key, sz_name in SIZE_KEY_MAP.items():
                ci = col_map.get(sz_key)
                if ci is not None and ci < len(row):
                    try: sizes[sz_name] = int(float(row[ci] or 0))
                    except: pass
            total_units = sum(sizes.values())
            # Fallback: if no size columns have values, use QTY column as OS (hats/caps)
            # Only use QTY as fallback if it comes BEFORE the size columns (not a total column)
            first_size_col = min((col_map[k] for k in SIZE_KEY_MAP if col_map.get(k) is not None), default=999)
            qty_col = col_map.get("qty")
            qty_is_before_sizes = qty_col is not None and qty_col < first_size_col
            if total_units == 0 and qty_is_before_sizes and qty_col < len(row):
                try:
                    qty_fallback = int(float(row[qty_col] or 0))
                    if qty_fallback > 0:
                        sizes["OS"] = qty_fallback
                        total_units = qty_fallback
                except: pass
            # Fallback: if all size cols are 0 but QTY col has a value, use as OS
            # Only apply for genuine new style rows (not carry-forward/total rows)
            if total_units == 0 and col_map.get("qty") is not None and col_map["qty"] < len(row):
                orig_style_val = ""
                if col_map.get("style") is not None and col_map["style"] < len(row):
                    orig_style_val = str(row[col_map["style"]] or "").strip()
                if orig_style_val and re.match(r"^[A-Za-z]{2,}\d+", orig_style_val):
                    try:
                        q = int(float(row[col_map["qty"]] or 0))
                        if q > 0:
                            sizes["OS"] = q
                            total_units = q
                    except: pass

        elif has_size_col and col_map.get("size") is not None:
            size_val = str(row[col_map["size"]] or "").strip()
            if col_map.get("qty") is not None and col_map["qty"] < len(row):
                try: qty_val = int(float(row[col_map["qty"]] or 0))
                except: pass
            if "/" in size_val:
                parts = size_val.split("/")
                size_val = parts[0].strip()
                if not color_code:
                    color_code = parts[1].strip()[:3]
            if size_val:
                norm = normalize_size(size_val)
                sizes[norm] = qty_val
                total_units = qty_val

        elif col_map.get("qty") is not None and col_map["qty"] < len(row):
            try:
                total_units = int(float(row[col_map["qty"]] or 0))
                sizes["OS"] = total_units
            except: pass

        if total_units == 0 and not desc:
            continue
        if total_units == 0 and not cost and not msrp:
            continue
        if has_size_col and not has_matrix and not size_val and not has_size_break:
            continue

        # If cost column is already net (COST W/DISC), don't apply discount again
        cost_is_net = col_map.get("cost_is_net", False)
        line_cost   = cost if cost_is_net else (cost * (1 - discount) if discount else cost)
        total_cost   = line_cost * total_units
        total_retail = msrp * total_units

        # Merge con línea existente (formato fila-por-talla)
        merged = False
        if has_size_col and size_val:
            for existing in reversed(lines):
                if existing["stock"] == stock and existing["color_code"] == color_code:
                    norm = normalize_size(size_val)
                    existing["sizes"][norm] = existing["sizes"].get(norm, 0) + qty_val
                    existing["total_units"] += qty_val
                    if not existing["cost"] and cost: existing["cost"] = cost
                    if not existing["msrp"] and msrp: existing["msrp"] = msrp
                    if not existing["discount"] and discount: existing["discount"] = discount
                    existing["line_cost"]    = existing["cost"] if col_map.get("cost_is_net") else (existing["cost"] * (1 - existing["discount"]) if existing["discount"] else existing["cost"])
                    existing["total_cost"]   = existing["line_cost"] * existing["total_units"]
                    existing["total_retail"] = existing["msrp"] * existing["total_units"]
                    merged = True
                    break

        if not merged:
            lines.append({
                "style_raw": style_raw, "stock": stock,
                "color_code": color_code, "color_name": color_name,
                "description": desc, "cost": cost, "msrp": msrp,
                "discount": discount, "line_cost": line_cost,
                "sizes": sizes, "total_units": total_units,
                "total_cost": total_cost, "total_retail": total_retail,
            })

    # ── Shiekh multi-sheet: parse remaining sheets ─────────────
    if is_shiekh and len(all_sheet_rows) > 1:
        for sheet_name, sheet_rows in all_sheet_rows[1:]:
            if sheet_name.upper() in ("TOTALS","SHEET1","SUMMARY"):
                continue
            # Find size header row (has OS, S, M, L, XL in row 5)
            shiekh_size_cols = {}
            for sr_idx, sr in enumerate(sheet_rows[:8]):
                sr_str = " | ".join(cell_str(c) for c in sr)
                if any(cell_str(c) in ("S","M","L","XL","OS","2XL","3XL") for c in sr if c):
                    for sc_idx, sc in enumerate(sr):
                        norm = normalize_size(cell_str(sc)) if sc else ""
                        if norm in SIZE_ORDER:
                            shiekh_size_cols[norm] = sc_idx
                    break
            if not shiekh_size_cols:
                continue
            # Get this sheet's PO# if different
            sheet_po = ""
            for sr in sheet_rows[:5]:
                if len(sr) > 18 and cell_str(sr[17]) == "PO#":
                    sheet_po = str(sr[18] or "").strip()
            # Find data row (style in col 3, desc in col 4, color in col 5)
            for sr in sheet_rows[6:]:
                style_raw = str(sr[3] or "").strip() if len(sr) > 3 else ""
                if not style_raw or not re.match(r"^[A-Za-z]{2,}\d+", style_raw):
                    continue
                desc  = str(sr[4] or "").strip() if len(sr) > 4 else ""
                color = str(sr[5] or "").strip() if len(sr) > 5 else ""
                cost  = float(sr[19] or 0) if len(sr) > 19 else 0
                msrp  = float(sr[20] or 0) if len(sr) > 20 else 0
                stock, cc, cn = parse_style_color(style_raw)
                if color and not cn: cn = color
                sizes = {s: 0 for s in SIZE_ORDER}
                for sz_name, sc_idx in shiekh_size_cols.items():
                    if sc_idx < len(sr):
                        try: sizes[sz_name] = int(float(sr[sc_idx] or 0))
                        except: pass
                total_units = sum(sizes.values())
                if total_units == 0: continue
                line_cost    = cost
                total_cost   = line_cost * total_units
                total_retail = msrp * total_units
                lines.append({
                    "style_raw": style_raw, "stock": stock,
                    "color_code": cc, "color_name": cn,
                    "description": desc, "cost": cost, "msrp": msrp,
                    "discount": 0, "line_cost": line_cost,
                    "sizes": sizes, "total_units": total_units,
                    "total_cost": total_cost, "total_retail": total_retail,
                })
                break  # one product per sheet in Shiekh format

    result["lines"] = lines
    return result

# ══════════════════════════════════════════════════════════════
# GENERAR ORDER FORM EXCEL
# ══════════════════════════════════════════════════════════════

def generate_order_form(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ORDER FORM"

    lf = Font(name="Calibri", bold=True, size=9, color="888888")
    vf = Font(name="Calibri", size=9)
    hf = Font(name="Calibri", bold=True, size=8, color="FFFFFF")
    bf = Font(name="Calibri", bold=True, size=9)

    totU = sum(l["total_units"] for l in data["lines"])
    totC = sum(l["total_cost"]  for l in data["lines"])

    def wlv(r, cl, label, cv, value):
        ws.cell(row=r, column=cl, value=label).font = lf
        ws.cell(row=r, column=cv, value=value).font = vf

    # Get discount % from first line that has one
    disc_pct = 0.0
    for l in data["lines"]:
        if l.get("discount"):
            disc_pct = l["discount"]
            break
    disc_label = f'{disc_pct*100:.0f}%' if disc_pct else ""
    cost_w_disc_header = f"COST W/DISCOUNT {disc_label}".strip()

    # Match ORDER_FORM.xlsx template exactly
    wlv(1,1,"Customer code:",2,data.get("customer_code",""))
    ws.cell(row=1,column=5,value="Vendor:").font=lf
    ws.cell(row=1,column=6,value="Maxima Apparel Corp SAPI de CV").font=vf
    ws.cell(row=1,column=8,value="Brand:").font=lf
    ws.cell(row=1,column=9,value="PRO STANDARD").font=bf

    wlv(3,1,"Customer name:",2,data.get("customer_name",""))
    wlv(3,5,"Ship to:",6,data.get("ship_to",""))
    wlv(5,1,"PO Number:",2,data.get("po_number",""))
    wlv(6,1,"PO Creation Date:",2,data.get("po_date",""))
    wlv(8,1,"Ship Date:",2,data.get("ship_date",""))
    wlv(9,1,"Cancel Date:",2,data.get("cancel_date",""))
    ws.cell(row=9,column=5,value="Total cost:").font=lf
    ws.cell(row=9,column=6,value=round(totC,2)).font=vf
    ws.cell(row=10,column=5,value="Total qty:").font=lf
    ws.cell(row=10,column=6,value=totU).font=vf
    ws.cell(row=11,column=5,value="Currency:").font=lf
    ws.cell(row=11,column=6,value=data.get("currency","USD")).font=vf
    ws.cell(row=12,column=5,value="Discount %:").font=lf
    ws.cell(row=12,column=6,value=disc_label).font=vf

    COLS = ["BRAND","STOCK","COLOR CODE","COLOR NAME","DESCRIPTION","CURRENCY",
            "LINE COST", cost_w_disc_header, "MSRP",
            "OS","XXS","XS","S","M","L","XL","2XL","3XL","4XL","5XL",
            "TOTAL UNITS","TOTAL COST","TOTAL RETAIL"]
    for j, cn in enumerate(COLS, start=1):
        c = ws.cell(row=13, column=j, value=cn)
        c.font = hf
        c.fill = PatternFill("solid", fgColor="222222")
        c.alignment = Alignment(horizontal="center")

    zero_fill = PatternFill("solid", fgColor="AAAAAA")   # gris para 0 unidades
    zero_font = Font(name="DM Sans", size=9, italic=True, color="666666")

    for i, line in enumerate(data["lines"], start=14):
        s = line["sizes"]
        rd = ["PRO STANDARD", line["stock"], line["color_code"], line["color_name"],
              line["description"], data.get("currency","USD"),
              round(line["cost"],2),        # LINE COST = costo bruto
              round(line["line_cost"],2),   # COST W/DISCOUNT = costo neto
              round(line["msrp"],2),
              s.get("OS",0), s.get("XXS",0), s.get("XS",0),
              s.get("S",0), s.get("M",0), s.get("L",0),
              s.get("XL",0), s.get("2XL",0), s.get("3XL",0),
              s.get("4XL",0), s.get("5XL",0),
              line["total_units"], round(line["total_cost"],2), round(line["total_retail"],2)]
        is_zero = line["total_units"] == 0
        for j, val in enumerate(rd, start=1):
            c = ws.cell(row=i, column=j, value=val)
            if is_zero:
                c.font = zero_font
                c.fill = zero_fill
            else:
                c.font = vf
            c.alignment = Alignment(horizontal="center" if j > 5 else "left")

    for i, w in enumerate([14,20,11,14,42,9,10,14,8,6,6,6,6,6,6,6,6,6,6,6,12,12,12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ══════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════

st.markdown('<p class="main-title">PURCHASE ORDER</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">PRO STANDARD — GENERATOR</p>', unsafe_allow_html=True)
st.divider()

col1, col2 = st.columns([1, 2])
with col1:
    st.markdown('<p class="section-title">📋 Orden del Cliente</p>', unsafe_allow_html=True)
    uploaded_file = st.file_uploader("Arrastra o selecciona", type=["xlsx","xls","pdf"], label_visibility="collapsed")

if uploaded_file:
    with st.spinner("Leyendo archivo..."):
        try:
            file_bytes = uploaded_file.read()
            ext = uploaded_file.name.split('.')[-1].lower()
            if ext == 'pdf':
                data = parse_po_pdf(file_bytes)
            else:
                data = parse_po_excel(file_bytes)
        except Exception as e:
            st.error(f"⚠️ Error al leer el archivo: {str(e)[:200]}")
            st.info("Intenta con otro archivo o contacta soporte.")
            st.stop()

    for w in data.get("warnings", []):
        st.markdown(f'<div class="warn-box">{w}</div>', unsafe_allow_html=True)

    # ── Diagnostic report ────────────────────────────────────
    diag = []
    diag.append(f"📄 Archivo: {uploaded_file.name}")
    diag.append(f"{'✅' if data['po_number'] else '❌'} PO Number: {data['po_number'] or 'NO DETECTADO'}")
    diag.append(f"{'✅' if data['customer_name'] else '❌'} Cliente: {data['customer_name'] or 'NO DETECTADO'}")
    diag.append(f"{'✅' if data['ship_date'] else '⚠️'} Ship Date: {data['ship_date'] or 'NO DETECTADO'}")
    diag.append(f"{'✅' if data['cancel_date'] else '⚠️'} Cancel Date: {data['cancel_date'] or 'NO DETECTADO'}")
    diag.append(f"{'✅' if data['lines'] else '❌'} Líneas de producto: {len(data['lines'])}")
    if data['lines']:
        zero_sizes = [l['stock'] for l in data['lines'] if l['total_units'] == 0]
        if zero_sizes:
            diag.append(f"⚠️  Estilos con 0 unidades: {', '.join(zero_sizes)}")
        no_desc = [l['stock'] for l in data['lines'] if not l['description']]
        if no_desc:
            diag.append(f"⚠️  Sin descripción: {', '.join(no_desc)}")
        no_cost = [l['stock'] for l in data['lines'] if l['cost'] == 0]
        if no_cost:
            diag.append(f"⚠️  Sin costo: {', '.join(no_cost)}")

    has_issues = any(l.startswith("❌") or l.startswith("⚠️") for l in diag)

    with st.expander("📋 Reporte de lectura" + (" ⚠️ hay campos sin detectar" if has_issues else " ✅ todo detectado"), expanded=has_issues):
        for line in diag:
            st.text(line)
        if has_issues:
            st.divider()
            st.markdown("**¿Algo salió mal?** Copia este reporte y mándalo para corregir el parser:")
            st.code("\n".join(diag))

    if data["lines"]:
        st.markdown('<div class="success-box">✓ ' + str(len(data["lines"])) + ' estilos detectados</div>', unsafe_allow_html=True)
        st.write("")

        st.markdown('<p class="section-title">Datos del Header</p>', unsafe_allow_html=True)
        hc = st.columns(4)
        for idx, (label, value) in enumerate([
            ("PO Number",data["po_number"]),("Cliente",data["customer_name"]),
            ("Ship Date",data["ship_date"]),("Cancel Date",data["cancel_date"]),
            ("PO Date",data["po_date"]),("Ship To",data["ship_to"]),
            ("Terms",data["terms"]),("Currency",data["currency"]),
        ]):
            with hc[idx % 4]:
                st.markdown(f'<div class="field-box"><div class="field-label">{label}</div><div class="field-value">{value or "—"}</div></div>', unsafe_allow_html=True)

        st.write("")
        st.markdown('<p class="section-title">Líneas de la Orden</p>', unsafe_allow_html=True)
        preview = []
        for l in data["lines"]:
            s = l["sizes"]
            preview.append({
                "STOCK":l["stock"],"COLOR":l["color_name"] or l["color_code"],
                "DESCRIPTION":l["description"],
                "OS":s.get("OS",0),"S":s.get("S",0),"M":s.get("M",0),
                "L":s.get("L",0),"XL":s.get("XL",0),"2XL":s.get("2XL",0),"3XL":s.get("3XL",0),
                "UNITS": f'0 ⚠️' if l["total_units"] == 0 else l["total_units"],
                "LINE COST":f'${l["line_cost"]:.2f}',"MSRP":f'${l["msrp"]:.2f}',
                "TOTAL COST":f'${l["total_cost"]:.2f}',
            })
        st.dataframe(pd.DataFrame(preview), width="stretch", hide_index=True)

        totU = sum(l["total_units"] for l in data["lines"])
        totC = sum(l["total_cost"]  for l in data["lines"])
        m1,m2,m3 = st.columns(3)
        m1.metric("Estilos", len(data["lines"]))
        m2.metric("Unidades", f"{totU:,}")
        m3.metric("Total Costo", f"${totC:,.2f}")

        st.write("")
        if st.button("⬇️  Generar ORDER FORM", type="primary", width="stretch"):
            with st.spinner("Generando..."):
                output   = generate_order_form(data)
                customer = (data["customer_name"] or data["po_number"] or "ORDER").replace("/","")[:30]
                filename = f"ORDER FORM - {customer}.xlsx"
            st.download_button("📥 Descargar ORDER FORM", data=output, file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch")
    else:
        st.markdown('<div class="warn-box">⚠️ No se detectaron líneas. Verifica el archivo.</div>', unsafe_allow_html=True)
else:
    with col1:
        st.markdown("""
        <div style="color:#555;font-size:0.85rem;line-height:1.8;margin-top:8px;">
        Formatos soportados:<br>
        • Tallas en filas (Kings / Caesars)<br>
        • Tallas en columnas (Fitters / Saleh)<br>
        • Size break comprimido (Pro Standard HO)<br>
        • Cualquier formato con encabezado en Excel
        </div>""", unsafe_allow_html=True)
